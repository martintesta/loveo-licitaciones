"""
consolidado.py — Tablero de decisión CONSOLIDADO de varias licitaciones, en un solo .xlsx.

El análisis integral (informe.py) responde "¿esta licitación se gana y a qué precio?". Esto responde
la otra pregunta: "de las N que tengo abiertas, ¿a cuáles le pongo la ficha?".

Sale 100% de la base (no llama a la IA, no cuesta tokens): cierre, techo, costo de la cubicación
preciada, score 6-D y alineación. Los TECHOS DE COSTO y el semáforo son FÓRMULAS: se edita el costo
estimado de una fila y su veredicto recalcula solo, con los mismos factores del modelo financiero
que usa el informe individual (analisis.FACTOR_SANO / FACTOR_LIMITE — una sola fuente).

  consolidado_xlsx(codigos=None, estado="nueva", solo_vigentes=True) -> bytes
  filas(codigos=None, ...)                                           -> dicts (lo que va al Excel)

  python consolidado.py                 # las vigentes (cierre >= hoy) sin descartar
  python consolidado.py 1234-5-LE26 ...
"""
import datetime
import io
import json

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

import analisis
import db
import scoring
from excel_analisis import (AMARILLO, AMBAR, AMBAR_T, BORDER, FMT_CLP, GRIS, NEGRO, ROJO, ROJO_T,
                            VERDE, VERDE_T, _anchos, _hdr, _title)


def filas(codigos=None, estado="nueva", solo_vigentes=True, hoy=None):
    """Una fila por licitación: lo que hace falta para decidir a cuál dedicarle el equipo.

    Por defecto solo las VIGENTES (cierre >= hoy): el tablero es para decidir dónde poner el equipo
    esta semana, y la base arrastra licitaciones descubiertas hace años que solo agregan ruido.
    Pasar códigos explícitos ignora el filtro (ahí el usuario ya eligió)."""
    hoy = hoy or datetime.date.today()
    with db.conn() as c:
        if codigos:
            marcas = ",".join("?" * len(codigos))
            rows = c.execute(f"SELECT * FROM licitaciones WHERE codigo IN ({marcas})",
                             tuple(codigos)).fetchall()
        elif solo_vigentes:
            rows = c.execute("SELECT * FROM licitaciones WHERE estado_revision=? "
                             "AND (admisible IS NULL OR admisible=1) "
                             "AND fecha_cierre IS NOT NULL AND fecha_cierre >= ? "
                             "ORDER BY fecha_cierre", (estado, hoy.isoformat())).fetchall()
        else:
            rows = c.execute("SELECT * FROM licitaciones WHERE estado_revision=? "
                             "AND (admisible IS NULL OR admisible=1) ORDER BY fecha_cierre",
                             (estado,)).fetchall()
        rows = [dict(r) for r in rows]
        out = []
        for lic in rows:
            cod = lic["codigo"]
            a = c.execute("SELECT presupuesto_clp, plazo_dias FROM analisis_bases WHERE codigo=? "
                          "ORDER BY id DESC LIMIT 1", (cod,)).fetchone()
            cub = c.execute(
                "SELECT SUM(ci.total) AS costo, COUNT(*) AS n, "
                "       SUM(CASE WHEN ci.total IS NULL THEN 1 ELSE 0 END) AS sin_precio "
                "FROM cubicacion_items ci JOIN cubicaciones cu ON cu.id=ci.cubicacion_id "
                "WHERE cu.codigo=?", (cod,)).fetchone()
            sj = {}
            if lic.get("score_json"):
                try:
                    sj = json.loads(lic["score_json"]) or {}
                except (ValueError, TypeError):
                    sj = {}
            dims = sj.get("dimensiones") or {}
            km, base = scoring._dist_min(lic.get("region"))
            techo = (a["presupuesto_clp"] if a and a["presupuesto_clp"] else lic.get("monto_estimado")) or 0
            faltan = []
            if not techo:
                faltan.append("presupuesto no publicado")
            if not (cub and cub["n"]):
                faltan.append("sin cubicación")
            elif cub["sin_precio"]:
                faltan.append(f"{cub['sin_precio']} partida(s) sin precio")
            if not a:
                faltan.append("bases sin analizar (Capa C)")
            out.append({
                "codigo": cod, "nombre": lic.get("nombre") or "", "organismo": lic.get("organismo") or "",
                "cierre": (lic.get("fecha_cierre") or "")[:10], "km": km or 0, "base": base or "",
                "techo": techo, "costo": (cub["costo"] if cub and cub["costo"] else 0),
                "score": lic.get("score_provisional") or 0,
                "triage": scoring.triage(lic.get("score_provisional") or 0),
                "alineacion": (dims.get("alineacion") or {}).get("score", ""),
                "plazo_dias": (a["plazo_dias"] if a else "") or "",
                "faltante": faltan,
            })
    out.sort(key=lambda x: (x["cierre"] or "9999", -x["score"]))
    return out


def _hoja_tablero(wb, datos, hoy):
    ws = wb.create_sheet("Tablero")
    _title(ws, f"LOTE DE LICITACIONES — TABLERO DE DECISIÓN · Loveo Construcciones · {hoy.isoformat()}", 12)
    ws.cell(2, 1, "Celdas amarillas = editables. El techo de costo y el semáforo son FÓRMULAS: "
                  "cambiá el costo estimado y el veredicto de esa fila recalcula solo.").font = Font(
        italic=True, size=9)

    par = 3
    for j, (k, v, fmt) in enumerate([("Factor costo SANO", analisis.FACTOR_SANO, "0.000"),
                                     ("Factor costo LÍMITE", analisis.FACTOR_LIMITE, "0.000"),
                                     ("Sobrecosto aplicado", analisis.PARAMS["sobrecosto"], "0%")]):
        ws.cell(par, 1 + j * 2, k).font = Font(bold=True)
        c = ws.cell(par, 2 + j * 2, v)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        c.border = BORDER

    cols = ["Cierre", "ID", "Organismo / Objeto", "km", "Techo neto", "Costo estimado",
            "Costo máx. SANO", "Costo máx. LÍMITE", "Semáforo", "Score", "Triage", "Aline."]
    r0 = 5
    for j, h in enumerate(cols, 1):
        _hdr(ws.cell(r0, j, h))

    r = r0 + 1
    for d in datos:
        ws.cell(r, 1, d["cierre"]).border = BORDER
        ws.cell(r, 2, d["codigo"]).border = BORDER
        ws.cell(r, 3, f'{d["organismo"]} — {d["nombre"][:70]}').border = BORDER
        ws.cell(r, 4, d["km"]).border = BORDER
        for col, val in ((5, d["techo"]), (6, d["costo"])):
            c = ws.cell(r, col, val)
            c.number_format = FMT_CLP
            c.border = BORDER
            if col == 6:
                c.fill = PatternFill("solid", fgColor=AMARILLO)     # editable
        for col, fx in ((7, f'=IF($E{r}=0,"",$E{r}/$B$3)'), (8, f'=IF($E{r}=0,"",$E{r}/$D$3)')):
            c = ws.cell(r, col, fx)
            c.number_format = FMT_CLP
            c.border = BORDER
        sem = (f'=IF(OR($E{r}=0,$F{r}=0),"INFO FALTANTE",'
               f'IF($F{r}<=$G{r},"SANO",IF($F{r}<=$H{r},"AJUSTADO","NO-GO")))')
        c = ws.cell(r, 9, sem)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        ws.cell(r, 10, d["score"]).border = BORDER
        ws.cell(r, 11, d["triage"]).border = BORDER
        ws.cell(r, 12, d["alineacion"]).border = BORDER
        r += 1

    if r > r0 + 1:
        rng = f"I{r0 + 1}:I{r - 1}"
        for txt, bg, fg in (("SANO", VERDE, VERDE_T), ("AJUSTADO", AMBAR, AMBAR_T),
                            ("NO-GO", ROJO, ROJO_T), ("INFO FALTANTE", GRIS, NEGRO)):
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=[f'"{txt}"'],
                fill=PatternFill("solid", bgColor=bg), font=Font(bold=True, color=fg)))
    n = ws.cell(r + 1, 1,
                f"Costo máx. SANO = techo neto / {analisis.FACTOR_SANO} → margen objetivo. "
                f"Costo máx. LÍMITE = techo neto / {analisis.FACTOR_LIMITE} → bajo eso el margen "
                f"mínimo del {analisis.PARAMS['margen_min']:.0%} es inalcanzable. Ambos factores ya "
                f"incluyen el sobrecosto del {analisis.PARAMS['sobrecosto']:.0%} y el financiamiento "
                f"por proyecto (misma fuente que el informe individual: analisis.PARAMS).")
    n.font = Font(italic=True, size=9)
    _anchos(ws, "ABCDEFGHIJKL", (11, 18, 60, 7, 15, 15, 15, 16, 15, 8, 12, 8))
    ws.freeze_panes = f"A{r0 + 1}"


def _hoja_faltante(wb, datos):
    ws = wb.create_sheet("Info faltante")
    _title(ws, "INFORMACIÓN FALTANTE — lo que impide cerrar cada análisis", 3)
    for j, h in enumerate(["ID", "Cierre", "Dato faltante"], 1):
        _hdr(ws.cell(3, j, h))
    r = 4
    for d in datos:
        for f in d["faltante"]:
            ws.cell(r, 1, d["codigo"]).border = BORDER
            ws.cell(r, 2, d["cierre"]).border = BORDER
            c = ws.cell(r, 3, f)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    if r == 4:
        ws.cell(r, 1, "Nada pendiente: todas las licitaciones del lote tienen techo y cubicación preciada.")
    _anchos(ws, "ABC", (18, 12, 90))
    ws.freeze_panes = "A4"


def _hoja_descartes(wb):
    """Por qué se descartó cada una. Sale de la tabla `descartes` — es el feedback real del motor
    de búsqueda, no una lista escrita a mano."""
    with db.conn() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT d.codigo, d.categoria, d.tipo, d.motivo, l.nombre, l.organismo "
            "FROM descartes d LEFT JOIN licitaciones l ON l.codigo = d.codigo "
            "ORDER BY d.ts DESC LIMIT 200").fetchall()]
    if not rows:
        return
    ws = wb.create_sheet("Fuera de alcance")
    _title(ws, "DESCARTADAS — falsos positivos del motor de búsqueda", 5)
    for j, h in enumerate(["ID", "Organismo", "Objeto", "Categoría / tipo", "Motivo"], 1):
        _hdr(ws.cell(3, j, h))
    r = 4
    for d in rows:
        vals = [d["codigo"], d["organismo"] or "", (d["nombre"] or "")[:90],
                " / ".join(x for x in [d["categoria"], d["tipo"]] if x), d["motivo"] or ""]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    n = ws.cell(r + 1, 1, "Si un patrón se repite acá, conviene ajustar las keywords del motor "
                          "para que no entre al pipeline (config_local.py).")
    n.font = Font(italic=True, size=9)
    _anchos(ws, "ABCDE", (18, 38, 60, 26, 60))
    ws.freeze_panes = "A4"


def consolidado_xlsx(codigos=None, estado="nueva", solo_vigentes=True, hoy=None):
    """Bytes del .xlsx consolidado. Devuelve {"error": ...} si no hay nada que consolidar."""
    hoy = hoy or datetime.date.today()
    datos = filas(codigos, estado, solo_vigentes=solo_vigentes, hoy=hoy)
    if not datos:
        return {"error": "No hay licitaciones que consolidar con ese filtro."}
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_tablero(wb, datos, hoy)
    _hoja_faltante(wb, datos)
    _hoja_descartes(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    import pathlib
    import sys
    cods = sys.argv[1:] or None
    r = consolidado_xlsx(cods)
    if isinstance(r, dict):
        print("ERROR:", r["error"])
        raise SystemExit(1)
    out = pathlib.Path(f"RESUMEN_LOTE_{datetime.date.today().isoformat()}.xlsx")
    out.write_bytes(r)
    print(f"OK: {out} ({len(filas(cods))} licitaciones, {len(r)} bytes)")

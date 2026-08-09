"""
excel_analisis.py — Arma el Excel del análisis integral (8-9 hojas) desde el dict de analisis.py.

Portado del generador de la skill `analizar-licitacion`, convertido de script a módulo (sin
sys.argv, sin estado global, sin escribir a disco: devuelve un Workbook).

EL PUNTO: todo lo derivado son FÓRMULAS ENCADENADAS, no valores congelados. Editar Cantidad o
Precio unitario en «Cubicación» recalcula, en este orden: costo base → sobrecosto → capital
adelantado → financiamiento → precios mínimo/objetivo → escenarios de oferta → P&L → KPIs del
Resumen. Lo único estático es la simulación Monte Carlo (corre en Python al generar), y la hoja
lo dice con un aviso en rojo.

Hojas: Resumen · Objeto y Alineación · Cubicación · Compras · Financiero · Riesgos · RACI ·
       Estrategia Ganar (si hay matriz de evaluación) · Competencia (si hay censo).

  construir(D) -> openpyxl.Workbook
"""
import random

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NEGRO = "1A1A1A"
AMARILLO = "F2C200"
GRIS = "F2F2F2"
VERDE, VERDE_T = "C6EFCE", "006100"
AMBAR, AMBAR_T = "FFEB9C", "9C6500"
ROJO, ROJO_T = "FFC7CE", "9C0006"
WHITE = "FFFFFF"
AZUL = "0070C0"

_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FMT_CLP = '"$"#,##0'
FMT_PCT = "0.0%"
FMT_X = '0.00"×"'

DEFAULTS = {"sobrecosto": 0.30, "markup_min": 1.25, "markup_obj": 1.45, "margen_min": 0.25,
            "tasa_financiamiento": 0.10, "ppm": 0.02, "impuesto_renta": 0.25,
            "estructura_mensual": 17_800_000}


# ----------------------------------------------------------------------- helpers de estilo
def _hdr(cell):
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=NEGRO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, text)
    c.font = Font(bold=True, size=14, color=NEGRO)
    c.fill = PatternFill("solid", fgColor=AMARILLO)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def money(v):
    try:
        return f"${v:,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return v


def pct(v):
    try:
        return f"{v * 100:.1f}%"
    except (TypeError, ValueError):
        return v


def _mcell(ws, row, col, value, bold=False, fill=None):
    c = ws.cell(row, col, value)
    c.number_format = FMT_CLP
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.border = BORDER
    return c


def _pcell(ws, row, col, value, bold=False, fill=None):
    c = ws.cell(row, col, value)
    c.number_format = FMT_PCT
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.border = BORDER
    return c


def _sem_fill(kind):
    return {"ok": (VERDE, VERDE_T), "warn": (AMBAR, AMBAR_T), "bad": (ROJO, ROJO_T)}.get(kind, (GRIS, NEGRO))


def _anchos(ws, letras, medidas):
    for col, w in zip(letras, medidas):
        ws.column_dimensions[col].width = w


def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ----------------------------------------------------------------------- modelo financiero
def _finanzas(D):
    """Los números que se calculan en Python (para semáforos y simulación). Las CELDAS del Excel
    no usan estos valores: usan fórmulas equivalentes. Los dos caminos tienen que dar lo mismo —
    `tests/test_excel_analisis.py` lo verifica sobre la cadena costo→precio."""
    P = dict(DEFAULTS)
    P.update(D.get("params") or {})
    techo = _f((D.get("presupuesto") or {}).get("monto_disponible"))
    costo_base = sum(_f(r.get("cantidad")) * _f(r.get("precio_unitario")) for r in D.get("cubicacion") or [])
    costo_ss = costo_base * (1 + P["sobrecosto"])
    cap = D.get("capital_adelantado")
    cap_adel = _f(cap) if cap else costo_ss
    costo_fin = cap_adel * P["tasa_financiamiento"]
    return {"P": P, "techo": techo, "costo_base": costo_base, "costo_ss": costo_ss,
            "cap_adel": cap_adel, "costo_fin": costo_fin,
            "precio_min": costo_ss * P["markup_min"], "precio_obj": costo_ss * P["markup_obj"],
            "cb_max_obj": techo / P["markup_obj"] / (1 + P["sobrecosto"]) if techo else 0,
            "cb_max_min": techo / P["markup_min"] / (1 + P["sobrecosto"]) if techo else 0}


def _margenes(fin, precio):
    P = fin["P"]
    ub = precio - fin["costo_ss"] - fin["costo_fin"]
    ppm = precio * P["ppm"]
    base = ub - ppm
    imp = max(0.0, base * P["impuesto_renta"])
    un = base - imp
    return {"precio": precio, "ub": ub, "mb": (ub / precio if precio else 0), "ppm": ppm,
            "base": base, "imp": imp, "un": un, "mn": (un / precio if precio else 0)}


def _semaforo(mn):
    return "ok" if mn >= 0.25 else ("warn" if mn >= 0.15 else "bad")


# ----------------------------------------------------------------------- estrategia (simulación)
def _escalones(c):
    """Normaliza un criterio no-precio a escalones con probabilidades que suman 1."""
    esc = c.get("escalones") or [{"pts": c.get("peso", 0), "cond": "Cumple", "p_rival": 0.5},
                                 {"pts": 0, "cond": "No cumple", "p_rival": 0.5}]
    esc = [dict(e) for e in esc]
    tot = sum(_f(e.get("p_rival")) for e in esc)
    if tot <= 0:
        for e in esc:
            e["p_rival"] = 1.0 / len(esc)
    elif abs(tot - 1.0) > 1e-9:
        for e in esc:
            e["p_rival"] = _f(e.get("p_rival")) / tot
    return esc


def _estrategia(D, fin):
    """Modelo formal + Monte Carlo. Devuelve None si no hay matriz de evaluación o no hay techo
    (sin presupuesto no hay precio relativo que simular)."""
    EV = D.get("evaluacion") or {}
    crits = EV.get("criterios") or []
    techo = fin["techo"]
    if not crits or not techo:
        return None
    c_precio = next((c for c in crits if c.get("tipo") == "continuo"), None)
    W = _f(c_precio.get("peso")) if c_precio else 0.0
    crits_np = [dict(c) for c in crits if c.get("tipo") != "continuo"]
    for c in crits_np:
        c["_esc"] = _escalones(c)
    if not crits_np:
        return None
    N_MAX = sum(max(e["pts"] for e in c["_esc"]) for c in crits_np)
    N_LOVEO = _f(EV.get("loveo_no_precio")) or N_MAX
    compuertas = [c for c in crits_np if c.get("es_compuerta")]

    def sobreprecio_max(nc):
        den = nc - (N_LOVEO - W)
        return None if den <= 0 else W / den - 1.0

    paso = 5 if N_MAX >= 40 else 2
    frontera, nc = [], N_MAX
    while nc >= max(0, N_MAX - 30):
        frontera.append((nc, sobreprecio_max(nc)))
        nc -= paso

    equiv = []
    for c in crits_np:
        pmax = max(e["pts"] for e in c["_esc"])
        perdidos = [pmax]
        distintos = sorted({e["pts"] for e in c["_esc"]}, reverse=True)
        if len(distintos) > 2:
            perdidos.append(pmax - distintos[1])
        for X in sorted(set(perdidos), reverse=True):
            if X <= 0:
                continue
            d = 1 - W / (W + X)
            equiv.append({"criterio": c["nombre"] + ("" if X == pmax else " (un escalón)"),
                          "pts": X, "desc": d, "costo": c.get("costo_loveo", "≈ $0")})
    equiv.sort(key=lambda x: -x["pts"])

    # Censo real: solo si TODOS los competidores traen su probabilidad técnica. Sin p_eett, θ=0 y
    # la simulación concluiría que nadie califica y Loveo gana siempre — una mentira peligrosa.
    COMPET = [dict(c) for c in (D.get("competencia") or {}).get("empresas") or []]
    usa_censo = bool(COMPET) and all(_f(c.get("p_eett")) > 0 for c in COMPET)
    for c in COMPET:
        c["_theta"] = (_f(c.get("p_eett")) * (_f(c.get("p_plazo")) or 1) * (_f(c.get("p_gar")) or 1)
                       * (_f(c.get("p_integ")) or 1) * (_f(c.get("p_form")) or 1))
    if usa_censo:
        theta_prom = sum(c["_theta"] for c in COMPET) / len(COMPET)
        p_ninguno = 1.0
        for c in COMPET:
            p_ninguno *= (1 - c["_theta"])
        e_rivales = sum(c["_theta"] for c in COMPET)
    else:
        theta_prom = p_ninguno = e_rivales = None

    SUP = EV.get("supuestos_competencia") or {}
    n_rng = SUP.get("n_competidores") or [2, 6]
    p_rng = SUP.get("precio_rival_pct") or [0.78, 1.00]
    N_SIM = int(SUP.get("n_simulaciones", 15000))

    def _sortear(esc, u):
        acc = 0.0
        for e in esc:
            acc += e["p_rival"]
            if u <= acc:
                return e["pts"]
        return esc[-1]["pts"]

    def simular(p_loveo, n_loveo, sup_n, sup_p, p_over=None, seed=20260808):
        rnd = random.Random(seed)
        wins = 0
        for _ in range(N_SIM):
            ofertas = []
            for _ in range(rnd.randint(int(sup_n[0]), int(sup_n[1]))):
                npz, fuera = 0, False
                for c in crits_np:
                    esc = c["_esc"]
                    if p_over and c["nombre"] in p_over:
                        pm = max(e["pts"] for e in esc)
                        distintos = sorted({e["pts"] for e in esc}, reverse=True)
                        got = pm if rnd.random() < p_over[c["nombre"]] else (
                            distintos[1] if len(distintos) > 1 else 0)
                    else:
                        got = _sortear(esc, rnd.random())
                    if c.get("es_compuerta") and got == 0:
                        fuera = True                      # 2 etapas: su sobre económico no se abre
                        break
                    npz += got
                if fuera:
                    continue
                ofertas.append((npz, rnd.uniform(sup_p[0], sup_p[1])))
            if not ofertas:
                wins += 1
                continue
            p_min = min([o[1] for o in ofertas] + [p_loveo])
            sc_l = W * (p_min / p_loveo) + n_loveo
            mejor = max(W * (p_min / pp) + nn for nn, pp in ofertas)
            if sc_l > mejor + 1e-12:
                wins += 1
            elif abs(sc_l - mejor) < 1e-9:                # desempate de bases: 1° oferta económica
                if p_loveo < min(pp for nn, pp in ofertas if abs(W * (p_min / pp) + nn - mejor) < 1e-9):
                    wins += 1
        return wins / N_SIM

    def simular_censo(p_loveo, n_loveo, sup_p, factor=1.0, seed=20260808):
        rnd = random.Random(seed)
        wins = 0
        for _ in range(N_SIM):
            ofertas = []
            for c in COMPET:
                if rnd.random() >= min(1.0, _f(c.get("p_eett")) * factor):
                    continue
                z = 25
                z += 20 if rnd.random() < (_f(c.get("p_plazo")) or 0.5) else (10 if rnd.random() < 0.6 else 0)
                z += 10 if rnd.random() < (_f(c.get("p_gar")) or 0.7) else 5
                if rnd.random() < (_f(c.get("p_integ")) or 0.6):
                    z += 5
                if rnd.random() < (_f(c.get("p_form")) or 0.85):
                    z += 5
                ofertas.append((z, rnd.uniform(sup_p[0], sup_p[1])))
            if not ofertas:
                wins += 1
                continue
            p_min = min([o[1] for o in ofertas] + [p_loveo])
            sc_l = W * (p_min / p_loveo) + n_loveo
            mejor = max(W * (p_min / pp) + nn for nn, pp in ofertas)
            if sc_l > mejor + 1e-12:
                wins += 1
        return wins / N_SIM

    GRID = [0.85, 0.90, 0.925, 0.95, 0.97, 0.98, 1.00]
    mc = []
    for p in GRID:
        pw = simular_censo(p, N_LOVEO, p_rng) if usa_censo else simular(p, N_LOVEO, n_rng, p_rng)
        m = _margenes(fin, techo * p)
        mc.append({"pct": p, "precio": techo * p, "pw": pw, "un": m["un"], "ve": pw * m["un"], "mn": m["mn"]})

    nom = compuertas[0]["nombre"] if compuertas else crits_np[0]["nombre"]
    G2 = [0.90, 0.925, 0.95, 0.97, 0.98, 1.00]
    ve_mat = {}
    if usa_censo:
        for nombre, fac, pp in [("Base (θ del censo)", 1.0, p_rng),
                                ("Rivales +50% más capaces", 1.5, p_rng),
                                ("Rivales 2× más capaces", 2.0, p_rng),
                                ("Rivales agresivos en precio", 1.0,
                                 [max(0.5, p_rng[0] - 0.12), p_rng[1] - 0.04]),
                                ("Todo en contra", 2.5, [max(0.5, p_rng[0] - 0.12), p_rng[1] - 0.04])]:
            ve_mat[nombre] = {p: simular_censo(p, N_LOVEO, pp, fac) * _margenes(fin, techo * p)["un"]
                              for p in G2}
    else:
        for nombre, nn, pp, ov in [
                ("Base (supuestos declarados)", n_rng, p_rng, None),
                ("Rivales muy capaces", n_rng, p_rng, {nom: 0.80}),
                ("Rivales agresivos en precio", n_rng, [max(0.5, p_rng[0] - 0.13), p_rng[1] - 0.05], None),
                ("Muchos oferentes", [n_rng[0] + 2, n_rng[1] + 4], p_rng, None),
                ("Nicho: pocos califican", n_rng, p_rng, {nom: 0.25}),
                ("Todo en contra", [n_rng[0] + 2, n_rng[1] + 4],
                 [max(0.5, p_rng[0] - 0.18), p_rng[1] - 0.10], {nom: 0.85})]:
            ve_mat[nombre] = {p: simular(p, N_LOVEO, nn, pp, ov) * _margenes(fin, techo * p)["un"]
                              for p in G2}
    regret = {p: max(max(v.values()) - v[p] for v in ve_mat.values()) for p in G2}
    p_robusto = min(G2, key=lambda p: regret[p])

    return {"W": W, "N_MAX": N_MAX, "N_LOVEO": N_LOVEO, "crits_np": crits_np, "c_precio": c_precio,
            "compuertas": compuertas, "frontera": frontera, "equiv": equiv, "mc": mc,
            "ve_mat": ve_mat, "regret": regret, "p_robusto": p_robusto, "n_rng": n_rng,
            "p_rng": p_rng, "N_SIM": N_SIM, "usa_censo": usa_censo, "compet": COMPET,
            "theta_prom": theta_prom, "p_ninguno": p_ninguno, "e_rivales": e_rivales, "EV": EV}


# ----------------------------------------------------------------------- hojas
def _hoja_resumen(wb, D, fin, dims, score_total, recom, penal, escenarios):
    ws = wb.active
    ws.title = "Resumen"
    meta = D["meta"]
    _title(ws, f"ANÁLISIS LICITACIÓN — {meta.get('id', '')} · {meta.get('nombre', '')}", 4)
    r = 3
    for k, v in [("Organismo", meta.get("organismo", "")),
                 ("Comuna / Región", f"{meta.get('comuna', '')} / {meta.get('region', '')}"),
                 ("Publicación / Cierre", f"{meta.get('fecha_publicacion', '')} → {meta.get('fecha_cierre', '')}"),
                 ("Días restantes", meta.get("dias_restantes", "")),
                 ("Base operativa / Distancia",
                  f"{meta.get('base_operativa', '')} · {int(_f(meta.get('distancia_km_base')))} km"),
                 ("Presupuesto disponible", money(fin["techo"]) if fin["techo"] else "NO PUBLICADO"),
                 ("Neto o IVA", (D.get("presupuesto") or {}).get("neto_o_iva", "desconocido")),
                 ("Fecha análisis", meta.get("fecha_analisis", ""))]:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1
    r += 1

    c = ws.cell(r, 1, "SCORING 6-D")
    c.font = Font(bold=True, size=12, color=NEGRO)
    c.fill = PatternFill("solid", fgColor=AMARILLO)
    r += 1
    for col, h in enumerate(["Dimensión", "Peso", "Puntaje (1-10)", "Subtotal"], 1):
        _hdr(ws.cell(r, col, h))
    r += 1
    primero = r
    for nombre, peso, sc in dims:
        ws.cell(r, 1, nombre)
        ws.cell(r, 2, peso)
        ws.cell(r, 3, sc)
        ws.cell(r, 4, f"=B{r}*C{r}")          # fórmula: editar el puntaje recalcula el total
        for col in range(1, 5):
            ws.cell(r, col).border = BORDER
        r += 1
    ws.cell(r, 1, "TOTAL / 120").font = Font(bold=True)
    ws.cell(r, 4, f"=SUM(D{primero}:D{r - 1})").font = Font(bold=True)
    r += 1
    rc = ws.cell(r, 1, f"Recomendación: {recom}")
    fg, ft = _sem_fill("ok" if score_total >= 70 else ("warn" if score_total >= 50 else "bad"))
    rc.fill = PatternFill("solid", fgColor=fg)
    rc.font = Font(bold=True, size=12, color=ft)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    ws.cell(r, 1, "SEMÁFORO FINANCIERO").font = Font(bold=True, size=12)
    r += 1
    if fin["costo_base"] and fin["techo"]:
        if fin["costo_base"] <= fin["cb_max_obj"]:
            sem, txt = "ok", "Sano — el costo entra bajo el techo de margen objetivo"
        elif fin["costo_base"] <= fin["cb_max_min"]:
            sem, txt = "warn", "Ajustado — entra, pero solo al margen mínimo"
        else:
            sem, txt = "bad", "Riesgo / NO-GO financiero — el costo supera el techo"
    else:
        sem, txt = "warn", "INFO FALTANTE (falta cubicación o presupuesto)"
    fg, ft = _sem_fill(sem)
    fc = ws.cell(r, 1, txt)
    fc.fill = PatternFill("solid", fgColor=fg)
    fc.font = Font(bold=True, color=ft)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    if penal:
        pc = ws.cell(r, 1, "⚠️ Penalización por distancia aplicada (>300 km): riesgo logístico ≤ 3")
        pc.font = Font(italic=True, color=ROJO_T)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1

    al = D.get("objeto_alineacion") or {}
    ws.cell(r, 1, "Alineación con Loveo (1-10)").font = Font(bold=True)
    ws.cell(r, 2, al.get("alineacion_score", ""))
    r += 1
    ws.cell(r, 1, al.get("alineacion_comentario", ""))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    vc = ws.cell(r, 1, f"🎯 VEREDICTO: {(D.get('estrategia') or {}).get('texto', '')}")
    vc.font = Font(bold=True)
    vc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 42
    r += 2

    c = ws.cell(r, 1, "NÚMEROS EN VIVO (se recalculan al editar la Cubicación)")
    c.font = Font(bold=True, size=12, color=NEGRO)
    c.fill = PatternFill("solid", fgColor=AMARILLO)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    kpi_row = r
    r += 8

    c = ws.cell(r, 1, "ESCENARIOS DE OFERTA vs. TECHO")
    c.font = Font(bold=True, size=12, color=NEGRO)
    c.fill = PatternFill("solid", fgColor=AMARILLO)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    esc_row = r
    r += 1 + len(escenarios) + 2

    ws.cell(r, 1, "INFO FALTANTE").font = Font(bold=True, color=ROJO_T)
    r += 1
    for it in D.get("info_faltante") or []:
        ws.cell(r, 1, f"• {it}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    _anchos(ws, "ABCDE", (44, 24, 16, 16, 14))
    return ws, kpi_row, esc_row


def _hoja_objeto(wb, D):
    ws = wb.create_sheet("Objeto y Alineación")
    _title(ws, "OBJETO Y ALINEACIÓN", 2)
    al = D.get("objeto_alineacion") or {}
    r = 3
    for k, v in [("Objeto", al.get("objeto", "")), ("Familia / tipo", al.get("familia", "")),
                 ("Alineación (1-10)", al.get("alineacion_score", "")),
                 ("Comentario de encaje", al.get("alineacion_comentario", "")),
                 ("Referencia competitiva", al.get("referencia_competitiva", ""))]:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90


def _hoja_cubicacion(wb, D):
    """La hoja EDITABLE. Todo el libro cuelga de la celda total de acá."""
    ws = wb.create_sheet("Cubicación")
    _title(ws, "CUBICACIÓN (Arquitecta)", 6)
    nota = ws.cell(2, 1, "✏️ EDITABLE: cambiá Cantidad o Precio unitario y TODO el análisis "
                         "(Financiero, escenarios, P&L, Resumen) se recalcula solo.")
    nota.font = Font(italic=True, size=9, color=AZUL)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    for col, h in enumerate(["Partida", "Unidad", "Cantidad", "Precio unitario", "Subtotal", "Notas"], 1):
        _hdr(ws.cell(3, col, h))

    r = 4
    primero = 4
    for it in D.get("cubicacion") or []:
        ws.cell(r, 1, it.get("partida", "")).border = BORDER
        ws.cell(r, 2, it.get("unidad", "")).border = BORDER
        ws.cell(r, 3, _f(it.get("cantidad"))).border = BORDER          # editable
        _mcell(ws, r, 4, _f(it.get("precio_unitario")))                # editable
        _mcell(ws, r, 5, f"=C{r}*D{r}")                                # fórmula
        nc = ws.cell(r, 6, it.get("notas", ""))
        nc.border = BORDER
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if r == primero:      # sin cubicación: una fila vacía editable, para que el rango exista y sirva
        ws.cell(r, 1, "(sin cubicación — agregá partidas acá)").border = BORDER
        ws.cell(r, 3, 0).border = BORDER
        _mcell(ws, r, 4, 0)
        _mcell(ws, r, 5, f"=C{r}*D{r}")
        ws.cell(r, 6, "").border = BORDER
        r += 1
    ultimo = r - 1

    ws.cell(r, 4, "COSTO BASE DIRECTO").font = Font(bold=True)
    _mcell(ws, r, 5, f"=SUM(E{primero}:E{ultimo})", bold=True, fill=AMARILLO)
    total_cell = f"'Cubicación'!$E${r}"
    r += 2
    if D.get("plazos"):
        ws.cell(r, 1, "PLAZOS DE EJECUCIÓN").font = Font(bold=True)
        r += 1
        for pl in D["plazos"]:
            ws.cell(r, 1, pl.get("etapa", ""))
            ws.cell(r, 2, f"{pl.get('dias', '')} días")
            r += 1
    _anchos(ws, "ABCDEF", (38, 10, 10, 18, 18, 34))
    ws.freeze_panes = "A4"
    return total_cell


def _hoja_compras(wb, D):
    ws = wb.create_sheet("Compras")
    _title(ws, "COMPRAS Y ABASTECIMIENTO", 6)
    for col, h in enumerate(["Ítem", "Proveedor candidato", "Lead time", "Precio ref.",
                             "Riesgo precio/stock", "Nota"], 1):
        _hdr(ws.cell(3, col, h))
    r = 4
    for it in D.get("compras") or []:
        for col, v in enumerate([it.get("item", ""), it.get("proveedor", ""), it.get("lead_time", ""),
                                 money(it.get("precio_ref", 0)), it.get("riesgo", ""),
                                 it.get("nota", "")], 1):
            c = ws.cell(r, col, v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if r == 4:
        ws.cell(r, 1, "Sin plan de compras: falta el análisis de bases con IA.").font = Font(
            italic=True, size=9, color=ROJO_T)
    _anchos(ws, "ABCDEF", (28, 26, 14, 16, 20, 30))


def _hoja_financiero(wb, D, fin, total_cell):
    """La cadena costo → precio, entera por fórmulas. Los parámetros (amarillo) son editables."""
    ws = wb.create_sheet("Financiero")
    _title(ws, "FINANCIERO (CFO)", 5)
    nota = ws.cell(2, 1, "✏️ Los parámetros (amarillo) son editables. Todo lo demás son fórmulas "
                         "encadenadas a la pestaña Cubicación — se recalcula solo.")
    nota.font = Font(italic=True, size=9, color=AZUL)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    P = fin["P"]
    r = 3
    ws.cell(r, 1, "PARÁMETROS (editables)").font = Font(bold=True)
    r += 1
    F = {}
    for key, label, val, fmt in [
            ("sobrecosto", "Sobrecosto", P["sobrecosto"], FMT_PCT),
            ("markup_min", "Markup precio mínimo", P["markup_min"], FMT_X),
            ("markup_obj", "Markup precio objetivo", P["markup_obj"], FMT_X),
            ("tasa_fin", "Tasa financiamiento (por proyecto)", P["tasa_financiamiento"], FMT_PCT),
            ("ppm", "PPM", P["ppm"], FMT_PCT),
            ("renta", "Impuesto renta", P["impuesto_renta"], FMT_PCT),
            ("margen_min", "Margen mínimo aceptable", P["margen_min"], FMT_PCT),
            ("estructura", "Estructura mensual", P["estructura_mensual"], FMT_CLP),
            ("techo", "TECHO / presupuesto disponible", fin["techo"], FMT_CLP)]:
        ws.cell(r, 1, label).font = Font(bold=True)
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        F[key] = f"$B${r}"
        r += 1
    r += 1

    ws.cell(r, 1, "CADENA COSTO → PRECIO").font = Font(bold=True)
    r += 1
    cap_fijo = D.get("capital_adelantado")
    for key, label in [("costo_base", "Costo base directo"),
                       ("costo_ss", "+ sobrecosto"),
                       ("cap_adel", "Capital adelantado"),
                       ("costo_fin", "Costo financiamiento (tasa × capital)"),
                       ("costo_total", "Costo total cargado"),
                       ("precio_min", "Precio mínimo"),
                       ("precio_obj", "Precio objetivo"),
                       ("cb_max_obj", "Costo base MÁX p/ margen objetivo"),
                       ("cb_max_min", "Costo base MÁX p/ margen mínimo")]:
        ws.cell(r, 1, label).font = Font(bold=True)
        F[key] = f"$B${r}"
        # if/elif y no un dict: cada fórmula referencia la celda de la fila ANTERIOR, así que las
        # ramas tienen que evaluarse de a una (un dict literal las evalúa todas y revienta).
        if key == "costo_base":
            val = f"={total_cell}"
        elif key == "costo_ss":
            val = f"={F['costo_base']}*(1+{F['sobrecosto']})"
        elif key == "cap_adel":
            val = _f(cap_fijo) if cap_fijo else f"={F['costo_ss']}"
        elif key == "costo_fin":
            val = f"={F['cap_adel']}*{F['tasa_fin']}"
        elif key == "costo_total":
            val = f"={F['costo_ss']}+{F['costo_fin']}"
        elif key == "precio_min":
            val = f"={F['costo_ss']}*{F['markup_min']}"
        elif key == "precio_obj":
            val = f"={F['costo_ss']}*{F['markup_obj']}"
        elif key == "cb_max_obj":
            val = f"={F['techo']}/{F['markup_obj']}/(1+{F['sobrecosto']})"
        else:
            val = f"={F['techo']}/{F['markup_min']}/(1+{F['sobrecosto']})"
        _mcell(ws, r, 2, val, bold=(key in ("costo_base", "precio_obj")),
               fill=(AMARILLO if key == "cap_adel" and cap_fijo else None))
        r += 1
    r += 1
    nc = ws.cell(r, 1, "➜ La TABLA DE ESCENARIOS DE OFERTA está en la pestaña «Resumen» "
                       "(es la tabla de decisión). Se alimenta de estas mismas celdas.")
    nc.font = Font(italic=True, bold=True, color=AZUL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    pct_rec = (D.get("estrategia") or {}).get("precio_recomendado_pct")
    if fin["techo"]:
        ws.cell(r, 1, "P&L AL PRECIO RECOMENDADO").font = Font(bold=True)
        ws.cell(r, 3, f"← editá el % en B{r + 1}").font = Font(italic=True, size=9, color=AZUL)
        r += 1
        ws.cell(r, 1, "% del techo a ofertar (editable)").font = Font(bold=True)
        c = ws.cell(r, 2, _f(pct_rec) or 0.95)
        c.number_format = FMT_PCT
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        c.border = BORDER
        F["pct_rec"] = f"$B${r}"
        r += 1
        ws.cell(r, 1, "Precio de oferta").font = Font(bold=True)
        _mcell(ws, r, 2, f"={F['techo']}*{F['pct_rec']}", bold=True)
        F["precio_rec"] = f"$B${r}"
        r += 1
        base_ref = util_ref = None
        for key, label in [("ingreso", "Ingreso (precio oferta)"), ("ppm", "− PPM"),
                           ("css", "− Costo con sobrecosto"), ("cfin", "− Financiamiento"),
                           ("base", "➜ Base imponible"), ("imp", "− Impuesto renta"),
                           ("util", "➜ UTILIDAD NETA"), ("rm", "Reparto Martin"),
                           ("rp", "Reparto Pablo")]:
            es_total = key in ("base", "util")
            ws.cell(r, 1, label).font = Font(bold=es_total)
            if key == "ingreso":
                val = f"={F['precio_rec']}"
            elif key == "ppm":
                val = f"=-{F['precio_rec']}*{F['ppm']}"
            elif key == "css":
                val = f"=-{F['costo_ss']}"
            elif key == "cfin":
                val = f"=-{F['costo_fin']}"
            elif key == "base":
                val = f"=SUM($B${r - 4}:$B${r - 1})"
                base_ref = f"$B${r}"
            elif key == "imp":
                val = f"=-MAX(0,{base_ref}*{F['renta']})"
            elif key == "util":
                val = f"={base_ref}+$B${r - 1}"
                util_ref = f"$B${r}"
            else:
                val = f"={util_ref}/2"
            _mcell(ws, r, 2, val, bold=es_total, fill=(AMARILLO if key == "util" else None))
            r += 1
        ws.cell(r, 1, "Margen neto %").font = Font(bold=True)
        _pcell(ws, r, 2, f"={util_ref}/{F['precio_rec']}", bold=True)
        r += 2
        ws.cell(r, 1, "Contribución a estructura mensual").font = Font(bold=True)
        cc = ws.cell(r, 2, f"={F['precio_rec']}/{F['estructura']}")
        cc.number_format = FMT_X
        cc.border = BORDER
        F["util_neta"] = util_ref
    _anchos(ws, "ABCDE", (40, 22, 16, 16, 16))
    return F


def _hoja_riesgos(wb, D):
    ws = wb.create_sheet("Riesgos")
    _title(ws, "RIESGOS Y MITIGACIÓN", 7)
    for col, h in enumerate(["Categoría", "Riesgo", "Prob.", "Impacto", "Mitigación",
                             "Costo mitigación", "Responsable"], 1):
        _hdr(ws.cell(3, col, h))
    r = 4
    primero = 4
    for it in D.get("riesgos") or []:
        for col, v in enumerate([it.get("categoria", ""), it.get("riesgo", ""), it.get("probabilidad", ""),
                                 it.get("impacto", ""), it.get("mitigacion", "")], 1):
            c = ws.cell(r, col, v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        _mcell(ws, r, 6, _f(it.get("costo_mitigacion")))          # editable
        ws.cell(r, 7, it.get("responsable", "")).border = BORDER
        r += 1
    if r == primero:      # sin riesgos, igual dejamos una fila para que el SUM tenga rango válido
        ws.cell(r, 1, "(sin riesgos cargados)").border = BORDER
        _mcell(ws, r, 6, 0)
        r += 1
    ws.cell(r, 5, "RESERVA DE CONTINGENCIA SUGERIDA").font = Font(bold=True)
    _mcell(ws, r, 6, f"=SUM(F{primero}:F{r - 1})", bold=True, fill=AMARILLO)
    _anchos(ws, "ABCDEFG", (20, 34, 8, 9, 40, 16, 14))


def _hoja_raci(wb, D):
    ws = wb.create_sheet("RACI")
    _title(ws, "MATRIZ RACI (R=Responsable · A=Aprueba · C=Consultado · I=Informado)", 4)
    for col, h in enumerate(["Ítem / Workstream", "Martin", "Pablo", "Valentina"], 1):
        _hdr(ws.cell(3, col, h))
    r = 4
    for it in D.get("raci") or []:
        for col, v in enumerate([it.get("item", ""), it.get("martin", ""), it.get("pablo", ""),
                                 it.get("valentina", "")], 1):
            c = ws.cell(r, col, v)
            c.border = BORDER
            if col > 1:
                c.alignment = Alignment(horizontal="center")
        r += 1
    ws.column_dimensions["A"].width = 42
    for col in "BCD":
        ws.column_dimensions[col].width = 12


def _hoja_estrategia(wb, D, fin, E, F):
    ws = wb.create_sheet("Estrategia Ganar")
    EV = E["EV"]
    q = [1]

    def t(txt, n=6):
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(bold=True, size=14, color=NEGRO)
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        ws.row_dimensions[q[0]].height = 26
        q[0] += 1

    def sec(txt, n=6):
        q[0] += 1
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NEGRO)
        ws.row_dimensions[q[0]].height = 20
        q[0] += 1

    def head(cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(q[0], i, h)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor="4A4A4A")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[q[0]].height = 30
        q[0] += 1

    def row(vals, fill=None, bold=False):
        for i, v in enumerate(vals, 1):
            c = ws.cell(q[0], i, v)
            c.border = BORDER
            c.font = Font(size=10, bold=bold)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        q[0] += 1

    def par(txt, n=6, fill=None, bold=False, h=None):
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(size=10, bold=bold)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            for i in range(1, n + 1):
                ws.cell(q[0], i).fill = PatternFill("solid", fgColor=fill)
        if h:
            ws.row_dimensions[q[0]].height = h
        q[0] += 1

    t(f"ESTRATEGIA PARA GANAR — {D['meta'].get('id', '')} · modelo formal + simulación")

    sec("1. ANATOMÍA DE LA MATRIZ — ¿dónde está realmente el puntaje?")
    head(["Criterio", "Peso", "Tipo de función", "Cómo se obtiene el máximo",
          "Costo marginal para Loveo", "Controlable"])
    if E["c_precio"]:
        cp = E["c_precio"]
        row([cp.get("nombre", "Oferta Económica"), cp.get("peso", ""),
             f"CONTINUA · {cp.get('formula') or '(P_min/P_i)×w'}", cp.get("como_max", ""),
             cp.get("costo_loveo") or "Margen — caro", cp.get("controlable") or "Parcial"], "DDEBF7")
    for c in E["crits_np"]:
        tipo = "BINARIA" if len(c["_esc"]) <= 2 else "ESCALONADA"
        det = " · ".join(str(e["pts"]) for e in sorted(c["_esc"], key=lambda x: -x["pts"]))
        row([c["nombre"] + ("  ⛔COMPUERTA" if c.get("es_compuerta") else ""),
             max(e["pts"] for e in c["_esc"]), f"{tipo} · {det}", c.get("como_max", ""),
             c.get("costo_loveo") or "≈ $0", c.get("controlable") or "TOTAL"], VERDE)
    row(["TOTAL NO-PRECIO", E["N_MAX"], "Binaria / escalonada",
         "Se asegura con disciplina, no con plata", "≈ $0", "TOTAL"], AMARILLO, bold=True)
    row(["TOTAL PRECIO", E["W"], "Continua", "Se compra con margen", "Alto", "Parcial"], AMBAR, bold=True)
    q[0] += 1
    if E["N_MAX"] > E["W"]:
        par(f"REGLA DE DECISIÓN: el puntaje no-precio ({E['N_MAX']}) SUPERA al de precio "
            f"({E['W']:.0f}) → esta licitación NO se gana por precio, se gana por ejecución "
            f"administrativa.", 6, AMARILLO, bold=True, h=30)
    else:
        par(f"REGLA DE DECISIÓN: el precio ({E['W']:.0f}) pesa más que el no-precio "
            f"({E['N_MAX']}) → acá SÍ es una pelea de precio. Verificar que el proyecto tenga "
            f"margen suficiente para pelearla.", 6, AMBAR, bold=True, h=30)

    if EV.get("no_evaluado"):
        sec("2. LO QUE **NO** SE EVALÚA — 0 puntos")
        par(" · ".join(EV["no_evaluado"]), 6, GRIS, h=28)
        if any("xperiencia" in str(x) for x in EV["no_evaluado"]):
            par("Que la EXPERIENCIA PREVIA pese 0 neutraliza la ventaja de los incumbentes grandes: "
                "es la razón estructural por la que Loveo compite de igual a igual. Además, "
                "sobre-especificar NO paga: si la calidad superior no da puntos, cumplir exacto es "
                "el óptimo económico.", 6, VERDE, h=44)

    if EV.get("reglas_estructurales"):
        sec("3. REGLAS ESTRUCTURALES QUE CAMBIAN EL JUEGO (fuera de la matriz)")
        head(["Artículo", "Regla", "Consecuencia estratégica", "", "", ""])
        col = {"critica": ROJO, "alta": AMBAR, "media": GRIS, "oportunidad": "DDEBF7"}
        for g in EV["reglas_estructurales"]:
            row([g.get("articulo", ""), g.get("regla", ""), g.get("consecuencia", ""), "", "", ""],
                col.get(g.get("severidad", "media"), GRIS))

    sec("4. MODELO FORMAL")
    par("PUNTAJE DE UN OFERENTE i:", 6, bold=True)
    par(f"    Score_i  =  {E['W']:.0f} · (P_min / P_i)  +  N_i        con N_i ∈ [0, {E['N_MAX']}]", 6, GRIS)
    par(f"FRONTERA DE INDIFERENCIA — Loveo (N={E['N_LOVEO']:.0f}) vence al rival C "
        f"(N_c, precio P_c) si:", 6, bold=True)
    par(f"    {E['W']:.0f} · (P_c/P_L) + {E['N_LOVEO']:.0f} > N_c + {E['W']:.0f}"
        f"        ⟺        P_L / P_c  <  {E['W']:.0f} / (N_c − {E['N_LOVEO'] - E['W']:.0f})", 6, GRIS)
    par(f"PRECIO EQUIVALENTE de X puntos no-precio:   descuento = 1 − {E['W']:.0f}/({E['W']:.0f}+X)",
        6, GRIS)

    sec("5. RESULTADO 1 — ¿Cuánto más caro puede ser Loveo y aún ganar?")
    head(["Puntaje no-precio del rival", "Sobreprecio máximo tolerable de Loveo", "Lectura", "", "", ""])
    for ncv, sp in E["frontera"]:
        if sp is None:
            row([ncv, "GANA SIEMPRE", "El rival no alcanza a Loveo con ningún precio", "", "", ""], VERDE)
        elif sp <= 0.001:
            row([ncv, "+0.0%", "Empate técnico → gana el más barato (desempate)", "", "", ""], ROJO)
        else:
            f = VERDE if sp >= 0.35 else (AMBAR if sp >= 0.10 else ROJO)
            row([ncv, f"+{sp * 100:.1f}%", f"Loveo gana aun siendo {sp * 100:.0f}% más caro", "", "", ""], f)

    sec("6. RESULTADO 2 — El Pareto: precio equivalente de cada punto no-precio")
    head(["Criterio que se pierde", "Puntos", "Descuento de precio equivalente",
          "En pesos sobre el techo", "Costo real de asegurarlo", ""])
    for e in E["equiv"]:
        f = VERDE if e["pts"] <= 10 else AMBAR
        row([e["criterio"], e["pts"], "", "", e["costo"], ""], f)
        rw = q[0] - 1
        _pcell(ws, rw, 3, e["desc"], fill=f)
        _mcell(ws, rw, 4, f"='Financiero'!{F['techo']}*$C${rw}", fill=f)   # vivo contra el techo

    sec("7. RESULTADO 3 — Simulación Monte Carlo (%s corridas por precio)"
        % f"{E['N_SIM']:,}".replace(",", "."))
    par("⚠️ ESTÁTICO: la simulación se corre en Python al generar el archivo. Si editás la "
        "Cubicación, las columnas de utilidad y valor esperado de ESTA tabla quedan desactualizadas "
        "— hay que regenerar el informe. Las secciones 1-6 y 9 sí son vivas.", 6, ROJO, bold=True, h=32)
    if E["usa_censo"]:
        par(f"✅ θ ESTIMADO DEL CENSO REAL. θ promedio = {E['theta_prom'] * 100:.1f}% · "
            f"P(ningún rival al máximo) = {E['p_ninguno'] * 100:.1f}% · "
            f"esperanza de rivales al máximo = {E['e_rivales']:.2f}. Ver pestaña «Competencia».",
            6, VERDE, bold=True, h=44)
    else:
        par("SUPUESTOS DECLARADOS (son hipótesis parametrizadas, NO datos observados):", 6, bold=True)
        sup = (f"nº competidores: {E['n_rng'][0]} a {E['n_rng'][1]}  ·  precio rival ∈ "
               f"[{E['p_rng'][0] * 100:.0f}%, {E['p_rng'][1] * 100:.0f}%] del techo  ·  ")
        sup += "  ·  ".join(f"P({c['nombre']} máx)="
                            f"{max(c['_esc'], key=lambda e: e['pts'])['p_rival'] * 100:.0f}%"
                            for c in E["crits_np"])
        par(sup, 6, GRIS, h=34)
    head(["% del techo", "Precio de oferta", "P(adjudicación)", "Utilidad neta si gana",
          "VALOR ESPERADO", "Margen neto"])
    ve_max = max(m["ve"] for m in E["mc"])
    for m in E["mc"]:
        f = VERDE if abs(m["ve"] - ve_max) < 1e-9 else None
        row([pct(m["pct"]), money(m["precio"]), pct(m["pw"]), money(m["un"]), money(m["ve"]),
             pct(m["mn"])], f, bold=(f == VERDE))

    sec("8. RESULTADO 4 — Prueba de robustez (minimax regret sobre %d escenarios)" % len(E["ve_mat"]))
    par("Se testeó el óptimo contra escenarios adversos. Se elige el precio que MINIMIZA el máximo "
        "arrepentimiento — es la decisión defendible, no el óptimo del caso base.", 6, GRIS, h=32)
    head(["% del techo", "Máximo arrepentimiento posible", "Veredicto", "", "", ""])
    reg_min = min(E["regret"].values())
    for p in sorted(E["regret"]):
        es_opt = abs(E["regret"][p] - reg_min) < 1e-9
        row([pct(p), money(E["regret"][p]), "ÓPTIMO ROBUSTO" if es_opt else "", "", "", ""],
            VERDE if es_opt else (AMBAR if E["regret"][p] < reg_min * 3 else ROJO), bold=es_opt)

    sec("9. TESIS Y DECISIÓN")
    if E["N_MAX"] > E["W"]:
        par("La pelea NO es de precio. Es de ADMISIBILIDAD y DISCIPLINA DOCUMENTAL.",
            6, AMARILLO, bold=True, h=22)
    p_rob = E["p_robusto"]
    pw_e = next((m["pw"] for m in E["mc"] if abs(m["pct"] - p_rob) < 1e-9), None)
    head(["Variable", "Valor", "", "", "", ""])
    row(["% del techo recomendado (robusto — editable)", "", "", "", "", ""], AMARILLO, bold=True)
    rw_pct = q[0] - 1
    _pcell(ws, rw_pct, 2, p_rob, bold=True, fill=AMARILLO)
    REF_PCT = f"$B${rw_pct}"
    row(["Precio de oferta recomendado", "", "", "", "", ""], VERDE, bold=True)
    rw_pr = q[0] - 1
    _mcell(ws, rw_pr, 2, f"='Financiero'!{F['techo']}*{REF_PCT}", bold=True, fill=VERDE)
    REF_PR = f"$B${rw_pr}"
    row(["Puntaje no-precio objetivo", f"{E['N_LOVEO']:.0f} / {E['N_MAX']} — innegociable",
         "", "", "", ""], AMARILLO, bold=True)
    row([f"Probabilidad estimada de adjudicación (@{pct(p_rob)}, estático)",
         pct(pw_e) if pw_e is not None else "—", "", "", "", ""])
    row(["Utilidad neta si gana (vivo)", "", "", "", "", ""])
    rw_u = q[0] - 1
    ub = f"({REF_PR}-'Financiero'!{F['costo_ss']}-'Financiero'!{F['costo_fin']})"
    _mcell(ws, rw_u, 2, f"=({ub}-{REF_PR}*'Financiero'!{F['ppm']})*(1-'Financiero'!{F['renta']})")
    row(["Margen neto (vivo)", "", "", "", "", ""])
    _pcell(ws, q[0] - 1, 2, f"=$B${rw_u}/{REF_PR}")
    row(["Contribución a estructura mensual (vivo)", "", "", "", "", ""])
    cc = ws.cell(q[0] - 1, 2, f"={REF_PR}/'Financiero'!{F['estructura']}")
    cc.number_format = FMT_X
    cc.border = BORDER
    par("⚠️ NO ofertar el techo matemático exacto: dejar holgura de redondeo del IVA. Verificar que "
        "precio_neto × 1,19 quede POR DEBAJO del techo con IVA incluido.", 6, GRIS, h=28)

    if EV.get("plan_accion"):
        sec("10. PLAN DE ACCIÓN ORDENADO POR RETORNO (Pareto)")
        head(["Prioridad", "Acción", "Puntos que asegura", "Costo", "Plazo", "Responsable"])
        for a in EV["plan_accion"]:
            pr = str(a.get("prioridad", ""))
            f = ROJO if any(x in pr for x in "012") else (AMBAR if any(x in pr for x in "345") else VERDE)
            row([pr, a.get("accion", ""), a.get("asegura", ""), a.get("costo", "≈ $0"),
                 a.get("plazo", ""), a.get("responsable", "")], f)

    _anchos(ws, "ABCDEF", (18, 46, 30, 22, 18, 18))
    ws.freeze_panes = "A2"


def _hoja_competencia(wb, D, E):
    CMP = D.get("competencia") or {}
    ws = wb.create_sheet("Competencia")
    q = [1]

    def t(txt, n=9):
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(bold=True, size=14, color=NEGRO)
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        ws.row_dimensions[q[0]].height = 26
        q[0] += 1

    def sec(txt, n=9):
        q[0] += 1
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NEGRO)
        ws.row_dimensions[q[0]].height = 20
        q[0] += 1

    def head(cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(q[0], i, h)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor="4A4A4A")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[q[0]].height = 32
        q[0] += 1

    def row(vals, fill=None, bold=False):
        for i, v in enumerate(vals, 1):
            c = ws.cell(q[0], i, v)
            c.border = BORDER
            c.font = Font(size=10, bold=bold)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        q[0] += 1

    def par(txt, n=9, fill=None, bold=False, h=None):
        ws.merge_cells(start_row=q[0], start_column=1, end_row=q[0], end_column=n)
        c = ws.cell(q[0], 1, txt)
        c.font = Font(size=10, bold=bold)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            for i in range(1, n + 1):
                ws.cell(q[0], i).fill = PatternFill("solid", fgColor=fill)
        if h:
            ws.row_dimensions[q[0]].height = h
        q[0] += 1

    t(f"INTELIGENCIA COMPETITIVA — {D['meta'].get('id', '')}")
    par(f"Fuente del censo: {CMP.get('fuente_censo', '—')}  ·  "
        f"Total de asistentes: {CMP.get('total_asistentes', '—')}", 9, GRIS, h=22)

    AM = {"ALTA": ROJO, "MEDIA-ALTA": AMBAR, "MEDIA": AMBAR, "MEDIA-BAJA": GRIS,
          "BAJA-MEDIA": GRIS, "BAJA": VERDE}
    sec("1. COMPETIDORES — perfil y amenaza")
    head(["Empresa", "Rubro real (giro SII)", "Antigüedad / tamaño", "Adjudicaciones históricas",
          "Historial CON este comprador", "Qué le FALTA para cumplir", "θ_i", "Amenaza", "Notas"])
    for c in CMP.get("empresas") or []:
        th = (_f(c.get("p_eett")) * (_f(c.get("p_plazo")) or 1) * (_f(c.get("p_gar")) or 1)
              * (_f(c.get("p_integ")) or 1) * (_f(c.get("p_form")) or 1))
        adj, monto = c.get("adjudicaciones_n"), c.get("adjudicaciones_monto")
        adjtxt = (f"{adj} lic. · {money(monto)}" if adj is not None and monto
                  else (f"{adj} licitaciones" if adj is not None else "—"))
        row([c.get("nombre", ""), c.get("rubro", ""), c.get("antiguedad", ""), adjtxt,
             c.get("historial_comprador", ""), c.get("le_falta", ""),
             (f"{th * 100:.1f}%" if th else "s/d"), c.get("amenaza", ""), c.get("nota", "")],
            AM.get(str(c.get("amenaza", "")).upper()))
    if E and E.get("usa_censo"):
        row(["AGREGADO", "", "", "", "", "", f"{E['theta_prom'] * 100:.1f}%", "θ promedio",
             f"P(ningún rival al máximo) = {E['p_ninguno'] * 100:.1f}% · "
             f"E[rivales al máximo] = {E['e_rivales']:.2f}"], AMARILLO, bold=True)
    else:
        par("θ sin estimar: el censo sale del historial de participaciones, no de un acta de visita "
            "con capacidades verificadas. Por eso la simulación de «Estrategia Ganar» corre con "
            "supuestos declarados y NO con este censo.", 9, GRIS, h=32)

    if CMP.get("benchmark_comprador"):
        sec("2. BENCHMARK DEL COMPRADOR — cómo adjudica esta unidad")
        head(["ID", "Objeto", "Presupuesto", "Oferentes", "Adjudicado", "% del ppto (c/IVA)",
              "Lectura", "", ""])
        for b in CMP["benchmark_comprador"]:
            pp, ad = _f(b.get("presupuesto")), b.get("adjudicado")
            pctv = f"{(_f(ad) * 1.19 / pp) * 100:.1f}%" if (ad and pp) else "—"
            row([b.get("id", ""), b.get("objeto", ""), money(pp), b.get("oferentes", "—"),
                 money(ad) if ad else "—", pctv, b.get("lectura", ""), "", ""],
                ROJO if b.get("resultado") == "DESIERTA" else None)

    if CMP.get("posicionamiento"):
        sec("3. POSICIONAMIENTO DE LOVEO")
        par(CMP["posicionamiento"], 9, VERDE, bold=True, h=60)

    _anchos(ws, "ABCDEFGHI", (28, 28, 20, 24, 30, 30, 9, 13, 40))
    ws.freeze_panes = "A2"


# ----------------------------------------------------------------------- entrada
def construir(D):
    """Dict de analisis.datos() -> Workbook con las 8-9 hojas."""
    fin = _finanzas(D)
    P = fin["P"]

    sc = D.get("scoring") or {}
    dist = _f((D.get("meta") or {}).get("distancia_km_base"))
    log = sc.get("logistico", 5)
    penal = dist > 300
    if penal:
        log = min(log, 3)
    dims = [("Margen potencial", 3, sc.get("margen", 5)), ("Cash flow timing", 3, sc.get("cashflow", 5)),
            ("Complejidad técnica", 2, sc.get("complejidad", 5)), ("Riesgo logístico", 2, log),
            ("Alineación capacidades", 1, sc.get("alineacion", 5)),
            ("Repetibilidad", 1, sc.get("repetibilidad", 5))]
    score_total = sum(w * s for _, w, s in dims)
    recom = ("GO FUERTE" if score_total >= 90 else "GO con mitigación" if score_total >= 70
             else "CONDICIONAL (solo si el pipeline está flojo)" if score_total >= 50 else "NO-GO")

    escenarios = []
    if fin["techo"]:
        for p in (0.80, 0.85, 0.90, 0.95, 1.00):
            m = _margenes(fin, fin["techo"] * p)
            m["pct"] = p
            m["sem"] = _semaforo(m["mn"])
            escenarios.append(m)

    wb = Workbook()
    ws_res, kpi_row, esc_row = _hoja_resumen(wb, D, fin, dims, score_total, recom, penal, escenarios)
    _hoja_objeto(wb, D)
    total_cell = _hoja_cubicacion(wb, D)
    _hoja_compras(wb, D)
    F = _hoja_financiero(wb, D, fin, total_cell)
    _hoja_riesgos(wb, D)
    _hoja_raci(wb, D)
    E = _estrategia(D, fin)
    if E:
        _hoja_estrategia(wb, D, fin, E, F)
    if (D.get("competencia") or {}).get("empresas"):
        _hoja_competencia(wb, D, E)

    # ---- pase final sobre el Resumen: ya existen las referencias de 'Financiero'.
    r = kpi_row
    kpis = [("Costo base directo (suma de la Cubicación)", f"='Financiero'!{F['costo_base']}", FMT_CLP),
            ("Costo con sobrecosto", f"='Financiero'!{F['costo_ss']}", FMT_CLP),
            ("Costo de financiamiento", f"='Financiero'!{F['costo_fin']}", FMT_CLP),
            ("Techo / presupuesto", f"='Financiero'!{F['techo']}", FMT_CLP)]
    if F.get("precio_rec"):
        kpis += [("Precio de oferta recomendado", f"='Financiero'!{F['precio_rec']}", FMT_CLP),
                 ("Utilidad neta si gana", f"='Financiero'!{F['util_neta']}", FMT_CLP),
                 ("Margen neto",
                  f"='Financiero'!{F['util_neta']}/'Financiero'!{F['precio_rec']}", FMT_PCT),
                 ("Contribución a estructura mensual",
                  f"='Financiero'!{F['precio_rec']}/'Financiero'!{F['estructura']}", FMT_X)]
    for lab, f, fmt in kpis:
        ws_res.cell(r, 1, lab).font = Font(bold=True)
        c = ws_res.cell(r, 2, f)
        c.number_format = fmt
        c.border = BORDER
        c.font = Font(bold=True)
        r += 1

    r = esc_row
    for col, h in enumerate(["% del techo", "Precio de oferta", "Margen bruto", "Margen neto",
                             "Semáforo"], 1):
        _hdr(ws_res.cell(r, col, h))
    r += 1
    primero = r
    FIN = "'Financiero'!"
    for e in escenarios:
        _pcell(ws_res, r, 1, e["pct"])
        _mcell(ws_res, r, 2, f"={FIN}{F['techo']}*$A${r}")
        ub = f"($B${r}-{FIN}{F['costo_ss']}-{FIN}{F['costo_fin']})"
        _pcell(ws_res, r, 3, f"={ub}/$B${r}")
        _pcell(ws_res, r, 4, f"=({ub}-$B${r}*{FIN}{F['ppm']})*(1-{FIN}{F['renta']})/$B${r}")
        sc_cell = ws_res.cell(r, 5, f'=IF($D${r}>={FIN}{F["margen_min"]},"OK",'
                                    f'IF($D${r}>=0.15,"AJUSTADO","RIESGO"))')
        sc_cell.font = Font(bold=True)
        sc_cell.alignment = Alignment(horizontal="center")
        sc_cell.border = BORDER
        r += 1
    ultimo = r - 1
    if ultimo >= primero:      # sin presupuesto publicado no hay escenarios → no hay rango que pintar
        rng = f"E{primero}:E{ultimo}"
        for txt, bg, fg in (("OK", VERDE, VERDE_T), ("AJUSTADO", AMBAR, AMBAR_T), ("RIESGO", ROJO, ROJO_T)):
            ws_res.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=[f'"{txt}"'],
                fill=PatternFill("solid", bgColor=bg), font=Font(bold=True, color=fg)))
        n = ws_res.cell(r, 1, f"Sweet spot = el precio más alto con margen neto ≥ {pct(P['margen_min'])}. "
                              f"Semáforo: OK ≥ {pct(P['margen_min'])} · "
                              f"AJUSTADO 15-{pct(P['margen_min'])} · RIESGO < 15%.")
    else:
        n = ws_res.cell(r, 1, "Sin presupuesto publicado: no hay techo contra el cual simular "
                              "escenarios de oferta. Cargalo en «Financiero» y la tabla se llena sola.")
    n.font = Font(italic=True, size=9, color=AZUL)
    ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    return wb

"""
informe.py — Exporta un .xlsx de una licitación con TODO lo que el sistema ya sabe.

Es deliberadamente de una sola fuente: no dispara research nuevo (eso es otro tipo de trabajo,
más caro y manual). Ensambla lo que el pipeline ya calculó — header, score 6-D, análisis de Capa C,
cubicación — en un Excel con FÓRMULAS reales (no valores pegados): el subtotal de cada partida y
el margen recalculan solos si el usuario edita cantidad o precio, igual que la plantilla de
referencia. Reusa scoring.PESOS/triage (misma fuente que el tablero) para que el informe nunca
pueda decir un número distinto al que muestra la UI.

  generar_xlsx(codigo) -> bytes (.xlsx) o {"error": ...} si el código no existe
"""
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import scoring
import cubicacion_ia

OXIDE = "B0501B"
STEEL = "6B7883"
HEADER_FILL = PatternFill("solid", fgColor="1C232B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color=OXIDE, bold=True, size=16)
SUB_FONT = Font(italic=True, size=11)
LABEL_FONT = Font(color=STEEL, bold=True, size=9.5)
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="D3DAE0")
BORDER = Border(bottom=THIN)

DIM_LABELS = {"margen": "Margen potencial", "cashflow": "Cash flow timing",
              "complejidad": "Complejidad técnica", "logistico": "Riesgo logístico",
              "alineacion": "Alineación", "repetibilidad": "Repetibilidad"}


def _lista(v):
    if isinstance(v, list):
        return v
    if isinstance(v, str) and v.strip():
        return [v.strip()]
    return []


def _header_row(ws, row, textos):
    for i, t in enumerate(textos, 1):
        cel = ws.cell(row=row, column=i, value=t)
        cel.font = HEADER_FONT
        cel.fill = HEADER_FILL
        cel.alignment = Alignment(vertical="center")


def _kv(ws, row, k, v):
    ws.cell(row=row, column=1, value=k).font = LABEL_FONT
    ws.cell(row=row, column=2, value=v)
    return row + 1


def _anchos(ws, anchos):
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a


def _recopilar(codigo):
    with db.conn() as c:
        lic = c.execute("SELECT * FROM licitaciones WHERE codigo=?", (codigo,)).fetchone()
        if not lic:
            return None
        lic = dict(lic)
        a = c.execute("SELECT * FROM analisis_bases WHERE codigo=? ORDER BY id DESC LIMIT 1",
                      (codigo,)).fetchone()
        analisis = dict(a) if a else None
        extraccion = {}
        if analisis and analisis.get("json_extraccion"):
            try:
                extraccion = json.loads(analisis["json_extraccion"])
            except (ValueError, TypeError):
                extraccion = {}
        items = cubicacion_ia.borrador(codigo)
    sj = {}
    if lic.get("score_json"):
        try:
            sj = json.loads(lic["score_json"])
        except (ValueError, TypeError):
            sj = {}
    return {"lic": lic, "analisis": analisis, "extraccion": extraccion, "items": items, "score_json": sj}


def _hoja_resumen(wb, codigo, lic, analisis, ex, sj):
    ws = wb.active
    ws.title = "Resumen"
    _anchos(ws, [26, 46, 14, 14])

    ws["A1"] = f"ANÁLISIS LICITACIÓN — {codigo}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = lic.get("nombre") or "—"
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    row = 4
    for k, v in [("Organismo", lic.get("organismo") or "—"),
                 ("Región", lic.get("region") or "—"),
                 ("Cierre", (lic.get("fecha_cierre") or "—")[:16]),
                 ("Presupuesto (portal)", lic.get("monto_estimado")),
                 ("Estado", f"{lic.get('estado_revision') or '—'} / {lic.get('estado_resultado') or 'pendiente'}"),
                 ("Admisible", "Sí" if lic.get("admisible") else ("No" if lic.get("admisible") is not None else "—"))]:
        row = _kv(ws, row, k, v)
    row += 1

    ws.cell(row=row, column=1, value="SCORING 6 DIMENSIONES").font = Font(bold=True, size=12)
    row += 1
    hdr_row = row
    _header_row(ws, row, ["Dimensión", "Peso", "Puntaje (1-10)", "Subtotal", "¿Evaluado?"])
    row += 1
    dims = sj.get("dimensiones") or {}
    primer_dato = row
    for k, peso in scoring.PESOS.items():
        dv = dims.get(k) or {}
        puntaje = dv.get("score")
        ws.cell(row=row, column=1, value=DIM_LABELS.get(k, k))
        ws.cell(row=row, column=2, value=peso)
        c_pje = ws.cell(row=row, column=3, value=puntaje if puntaje is not None else 0)
        # fórmula real: Subtotal = Peso × Puntaje — igual que la plantilla de referencia
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        ws.cell(row=row, column=5, value="Sí" if dv.get("evaluado") else "Pendiente")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    ultimo_dato = row - 1
    ws.cell(row=row, column=1, value="TOTAL / 120").font = TOTAL_FONT
    tot_cell = ws.cell(row=row, column=4, value=f"=SUM(D{primer_dato}:D{ultimo_dato})")
    tot_cell.font = TOTAL_FONT
    row += 2

    total_hoy = sj.get("score_provisional")
    ws.cell(row=row, column=1, value="Recomendación (según score al generar)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=scoring.triage(total_hoy) if total_hoy is not None else "sin score")
    row += 2

    if analisis:
        ws.cell(row=row, column=1, value="ANÁLISIS DE BASES (CAPA C)").font = Font(bold=True, size=12)
        row += 1
        if analisis.get("resumen"):
            ws.cell(row=row, column=1, value=analisis["resumen"])
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.row_dimensions[row].height = 30
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        for k, v in [("Presupuesto (bases)", analisis.get("presupuesto_clp")),
                     ("Plazo de entrega", f"{analisis['plazo_dias']} días" if analisis.get("plazo_dias") else "—"),
                     ("Subcontratación prohibida", str(ex.get("subcontratacion_prohibida") or "—")),
                     ("Anticipo", str(ex.get("anticipo") or "—"))]:
            row = _kv(ws, row, k, v)
        row += 1
        reqs = _lista(ex.get("requisitos_clave"))
        if reqs:
            ws.cell(row=row, column=1, value="Requisitos clave (checklist)").font = Font(bold=True, size=10.5)
            row += 1
            for x in reqs:
                ws.cell(row=row, column=1, value="☐")
                ws.cell(row=row, column=2, value=str(x))
                row += 1
            row += 1
        gars = _lista(ex.get("garantias"))
        if gars:
            ws.cell(row=row, column=1, value="Garantías").font = Font(bold=True, size=10.5)
            row += 1
            for g in gars:
                txt = f"{g.get('tipo', '')} — {g.get('monto_o_pct', '')}" if isinstance(g, dict) else str(g)
                ws.cell(row=row, column=1, value="☐")
                ws.cell(row=row, column=2, value=txt)
                row += 1
            row += 1
        crits = _lista(ex.get("criterios_evaluacion"))
        if crits:
            ws.cell(row=row, column=1, value="Criterios de evaluación").font = Font(bold=True, size=10.5)
            row += 1
            for cr in crits:
                if isinstance(cr, dict):
                    ws.cell(row=row, column=1, value=cr.get("nombre", ""))
                    ws.cell(row=row, column=2, value=f"{cr.get('peso_pct', '')}%")
                else:
                    ws.cell(row=row, column=1, value=str(cr))
                row += 1

    ws.freeze_panes = "A4"
    return hdr_row  # por si algún caller quiere referenciarlo (no usado hoy)


def _hoja_cubicacion(wb, codigo, items, analisis):
    ws = wb.create_sheet("Cubicación")
    _anchos(ws, [10, 42, 11, 9, 15, 15])
    ws["A1"] = "CUBICACIÓN"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "✏️ Editable: cambiá Cantidad o Precio unitario y el Total de esa fila recalcula solo."
    ws["A2"].font = Font(italic=True, size=9, color=STEEL)
    ws.merge_cells("A2:F2")

    hdr = 4
    _header_row(ws, hdr, ["Partida", "Descripción", "Cantidad", "Unidad", "Precio unitario", "Total"])
    row = hdr + 1
    primer = row
    if items:
        for it in items:
            ws.cell(row=row, column=1, value=it.get("partida") or "")
            ws.cell(row=row, column=2, value=it.get("descripcion") or "")
            ws.cell(row=row, column=3, value=it.get("cantidad"))
            ws.cell(row=row, column=4, value=it.get("unidad") or "")
            ws.cell(row=row, column=5, value=it.get("precio_unitario"))
            # fórmula: Total = Cantidad × Precio unitario (recalcula si se editan)
            ws.cell(row=row, column=6, value=f"=IF(AND(C{row}<>\"\",E{row}<>\"\"),C{row}*E{row},\"\")")
            ws.cell(row=row, column=5).number_format = "#,##0"
            ws.cell(row=row, column=6).number_format = "#,##0"
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Sin borrador de cubicación todavía.")
        row += 1
    ultimo = row - 1

    row += 1
    ws.cell(row=row, column=5, value="Costo base (Σ Total)").font = TOTAL_FONT
    costo_row = row
    ws.cell(row=row, column=6, value=(f"=SUM(F{primer}:F{ultimo})" if items else 0)).font = TOTAL_FONT
    ws.cell(row=row, column=6).number_format = "#,##0"
    row += 1

    techo = analisis.get("presupuesto_clp") if analisis else None
    if techo:
        ws.cell(row=row, column=5, value="Techo / presupuesto (bases)").font = LABEL_FONT
        ws.cell(row=row, column=6, value=techo).number_format = "#,##0"
        techo_row = row
        row += 1
        ws.cell(row=row, column=5, value="Margen (techo − costo)").font = TOTAL_FONT
        ws.cell(row=row, column=6, value=f"=F{techo_row}-F{costo_row}").number_format = "#,##0"
        margen_row = row
        row += 1
        ws.cell(row=row, column=5, value="Margen %").font = LABEL_FONT
        ws.cell(row=row, column=6, value=f"=IF(F{techo_row}=0,\"\",F{margen_row}/F{techo_row})")
        ws.cell(row=row, column=6).number_format = "0.0%"
        row += 1
    else:
        ws.cell(row=row, column=5, value="Sin techo de presupuesto (bases): falta Capa C.").font = Font(
            italic=True, size=9, color=STEEL)
        row += 1

    sin_precio = sum(1 for it in items if it.get("total") is None) if items else 0
    if sin_precio:
        ws.cell(row=row, column=1,
               value=f"⚠ {sin_precio} partida(s) sin precio o cantidad — el margen no cierra hasta completarlas.")
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color=STEEL)

    ws.freeze_panes = f"A{hdr + 1}"


def generar_xlsx(codigo):
    """Arma el .xlsx en memoria. Devuelve bytes, o {"error": ...} si el código no existe."""
    d = _recopilar(codigo)
    if d is None:
        return {"error": f"No se encontró la licitación {codigo}."}
    lic, analisis, ex, items, sj = d["lic"], d["analisis"], d["extraccion"], d["items"], d["score_json"]

    wb = Workbook()
    _hoja_resumen(wb, codigo, lic, analisis, ex, sj)
    _hoja_cubicacion(wb, codigo, items, analisis)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

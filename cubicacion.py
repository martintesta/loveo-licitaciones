"""
cubicacion.py — Ingesta el Excel de cubicación de Valentina a la base (schema_v3).

Tolerante al formato (puede variar): detecta los ingresos, las partidas (col B) y los ítems
(col C, o col B cuando la partida es un ítem único como TRASLADO/SILLON). De cada ítem saca
descripción, cantidad, unidad, precio unitario, total, devengado y estado de pago, y alimenta
también `precios_referencia` (catálogo para la cubicación tipo + trazabilidad de costos).

  python cubicacion.py "<ruta.xlsx>" [--proyecto NOMBRE] [--codigo COD]
  python cubicacion.py "<ruta.xlsx>" --dry      # no escribe, solo muestra el resumen

Estructura observada (Talagante): arriba "Ingreso con IVA/sin IVA"; filas: B=partida, C=ítem,
D=cantidad, E=unidad, F=precio unit, G=total, H=devengado, I=estado pago; abajo "Costo Total".
"""
import os
import sys
import unicodedata
import datetime
import openpyxl
import db

STOP = {
    "ingreso con iva", "ingreso sin iva", "costo total", "falta gastar",
    "presupuesto devengado", "total prespuesto", "total presupuesto",
    "estado de pagos/ fecha", "estado de pagos / fecha", "total",
}


def _na(s):
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower().strip()


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _grupo(partida, desc):
    p = _na(partida)
    t = p + " " + _na(desc)
    if "flete" in t or "traslado" in t:
        return "transporte"
    if "mano de obra" in p or "maestro" in t or "te1" in t or p == "instalacion":
        return "mano_obra"
    if p in ("otros", "sillon") or "sillon" in t:
        return "otros"
    return "material"


def _labeled(rows, label):
    """Primer valor numérico que sigue a una celda igual a `label` (ej. 'Costo Total')."""
    lab = _na(label)
    for row in rows:
        for i, v in enumerate(row):
            if _na(v) == lab:
                for v2 in row[i + 1:]:
                    n = _num(v2)
                    if n is not None:
                        return n
    return None


def _infer_proyecto(path):
    d = os.path.dirname(os.path.abspath(path))
    parent = os.path.basename(d)
    if _na(parent) in ("costos", "costo"):
        parent = os.path.basename(os.path.dirname(d))
    return parent or os.path.splitext(os.path.basename(path))[0]


def _version(path):
    return os.path.splitext(os.path.basename(path))[0]


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]

    ingreso_civa = _labeled(rows, "Ingreso con IVA")
    ingreso_siva = _labeled(rows, "Ingreso sin IVA")
    costo_total_file = _labeled(rows, "Costo Total")

    items = []
    partida = None
    for row in rows:
        def c(i):
            return row[i - 1] if i - 1 < len(row) else None
        B, C, D, E, F, G, H, I = c(2), c(3), c(4), c(5), c(6), c(7), c(8), c(9)

        if B not in (None, "") and _na(B) not in STOP:
            partida = str(B).strip()

        desc = str(C).strip() if C not in (None, "") else None
        cant, pu, tot = _num(D), _num(F), _num(G)

        if (pu is not None or tot is not None) and (desc or cant is not None):
            d = desc or (partida or "—")
            if _na(d) in STOP:
                continue
            items.append({
                "partida": partida or "", "grupo": _grupo(partida, d), "descripcion": d,
                "cantidad": cant, "unidad": (str(E).strip() if E not in (None, "") else None),
                "precio_unitario": pu, "total": tot, "devengado": _num(H),
                "estado_pago": (str(I).strip() if (I not in (None, "") and _num(I) is None) else None),
            })

    costo_total = costo_total_file if costo_total_file is not None else sum((it["total"] or 0) for it in items)
    margen = (ingreso_siva - costo_total) if ingreso_siva is not None else None
    margen_pct = (margen / ingreso_siva * 100) if (margen is not None and ingreso_siva) else None
    return {
        "ingreso_con_iva": ingreso_civa, "ingreso_sin_iva": ingreso_siva,
        "costo_total": costo_total, "margen": margen, "margen_pct": margen_pct, "items": items,
    }


def ingest(path, proyecto=None, codigo=None):
    data = parse(path)
    proyecto = proyecto or _infer_proyecto(path)
    archivo = os.path.basename(path)
    fecha = datetime.date.today().isoformat()
    items = data["items"]

    with db.conn() as conn:
        conn.execute("DELETE FROM cubicaciones WHERE proyecto=? AND archivo=?", (proyecto, archivo))
        conn.execute("DELETE FROM precios_referencia WHERE proyecto=? AND fuente='valentina'", (proyecto,))
        cur = conn.execute("""
            INSERT INTO cubicaciones(proyecto, codigo, archivo, ingreso_con_iva, ingreso_sin_iva,
                                     costo_total, margen, margen_pct, origen, fecha, version)
            VALUES (?,?,?,?,?,?,?,?, 'valentina', ?, ?) RETURNING id""",
            (proyecto, codigo, archivo, data["ingreso_con_iva"], data["ingreso_sin_iva"],
             data["costo_total"], data["margen"], data["margen_pct"], fecha, _version(path)))
        cid = cur.fetchone()[0]
        precios = 0
        for it in items:
            conn.execute("""
                INSERT INTO cubicacion_items(cubicacion_id, partida, grupo, descripcion, cantidad,
                                             unidad, precio_unitario, total, devengado, estado_pago)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (cid, it["partida"], it["grupo"], it["descripcion"], it["cantidad"], it["unidad"],
                 it["precio_unitario"], it["total"], it["devengado"], it["estado_pago"]))
            if it["precio_unitario"] is not None and it["descripcion"]:
                conn.execute("""
                    INSERT INTO precios_referencia(descripcion, descripcion_norm, unidad, precio_unitario,
                                                   fuente, proyecto, region, fecha)
                    VALUES (?,?,?,?, 'valentina', ?, NULL, ?)""",
                    (it["descripcion"], _na(it["descripcion"]), it["unidad"], it["precio_unitario"], proyecto, fecha))
                precios += 1

    return {"proyecto": proyecto, "archivo": archivo, "items": len(items), "precios": precios,
            "ingreso_sin_iva": data["ingreso_sin_iva"], "costo_total": data["costo_total"],
            "margen": data["margen"], "margen_pct": data["margen_pct"]}


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print("Uso: python cubicacion.py \"<ruta.xlsx>\" [--proyecto NOMBRE] [--codigo COD] [--dry]")
        sys.exit(1)
    path = args[0]
    proyecto = args[args.index("--proyecto") + 1] if "--proyecto" in args else None
    codigo = args[args.index("--codigo") + 1] if "--codigo" in args else None
    db.init_db()
    schema = __import__("schema_v3"); schema.main()
    if "--dry" in args:
        import json
        d = parse(path)
        d_resumen = {k: v for k, v in d.items() if k != "items"}
        d_resumen["n_items"] = len(d["items"])
        print(json.dumps(d_resumen, ensure_ascii=False, indent=2))
    else:
        import json
        print(json.dumps(ingest(path, proyecto, codigo), ensure_ascii=False, indent=2))

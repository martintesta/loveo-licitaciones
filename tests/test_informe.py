"""
test_informe.py — Sensor del ANÁLISIS INTEGRAL (informe.py + analisis.py + excel_analisis.py).

El informe pasó de 2 hojas a 8-9: Resumen, Objeto y Alineación, Cubicación, Compras, Financiero,
Riesgos, RACI, y — cuando hay matriz de evaluación / censo — Estrategia Ganar y Competencia.

Lo que estos tests protegen:
  1. Genera aunque falte todo (la mayoría de las licitaciones reales están incompletas).
  2. NUNCA llama a la red: `enriquecer` es inyectable y acá se inyecta siempre.
  3. Las celdas derivadas son FÓRMULAS, no valores pegados (el punto del Excel sobre el Word).
  4. Se guarda SIEMPRE en la carpeta de bases de la licitación (requisito explícito).
  5. El veredicto nunca contradice a scoring.triage (misma fuente que el tablero).

  python -m pytest tests/test_informe.py -q
"""
import io
import json

import pytest
from openpyxl import load_workbook

import analisis
import db
import informe
import schema_v2
import schema_v3
import scoring


def _setup(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DATABASE_URL", "")
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr(db, "STORAGE_DIR", tmp_path / "storage")
    monkeypatch.delenv("LOVEO_BASES_LOCAL", raising=False)
    db.init_db()
    schema_v2.migrate()      # competidores/participaciones/adjudicaciones (censo de competencia)
    schema_v3.migrate()


def _sin_ia(*a, **k):
    """`enriquecer` que devuelve nada: simula 'no hay API key'. El análisis debe seguir en pie."""
    return {}


def _con_ia(*a, **k):
    return {
        "objeto_alineacion": {"objeto": "2 containers", "familia": "contenedor", "alineacion_score": 9,
                              "alineacion_comentario": "Core de Loveo", "referencia_competitiva": "Talagante"},
        "cubicacion": [], "plazos": [{"etapa": "Fabricación", "dias": 30}],
        "compras": [{"item": "Container 20ft", "proveedor": "Prov X", "lead_time": "2 sem",
                     "precio_ref": 2200000, "riesgo": "Medio", "nota": ""}],
        "riesgos": [{"categoria": "Financiero", "riesgo": "Pago a 60 días", "probabilidad": "Alta",
                     "impacto": "Alto", "mitigacion": "Línea de capital", "costo_mitigacion": 350000,
                     "responsable": "Martin"}],
        "raci": [{"item": "GO/NO-GO", "martin": "A", "pablo": "C", "valentina": "C"}],
        "estrategia": {"texto": "Se gana por disciplina documental.", "precio_recomendado_pct": 0.95},
        "info_faltante": ["Monto neto vs IVA"],
        "evaluacion": {
            "criterios": [
                {"nombre": "Oferta Económica", "peso": 40, "tipo": "continuo", "formula": "(P_min/P_i)x40",
                 "como_max": "Ser el más barato", "costo_loveo": "Margen", "controlable": "Parcial",
                 "es_compuerta": False, "escalones": []},
                {"nombre": "EETT", "peso": 35, "tipo": "binario", "formula": "", "como_max": "Cumplir Anexo A",
                 "costo_loveo": "≈ $0", "controlable": "TOTAL", "es_compuerta": True,
                 "escalones": [{"pts": 35, "cond": "Cumple", "p_rival": 0.45},
                               {"pts": 0, "cond": "No cumple", "p_rival": 0.55}]},
                {"nombre": "Plazo", "peso": 25, "tipo": "escalonado", "formula": "", "como_max": "≤30 días",
                 "costo_loveo": "Stock", "controlable": "TOTAL",
                 "escalones": [{"pts": 25, "cond": "≤30d", "p_rival": 0.4},
                               {"pts": 12, "cond": "31-40d", "p_rival": 0.4},
                               {"pts": 0, "cond": ">40d", "p_rival": 0.2}]}],
            "no_evaluado": ["Experiencia previa"],
            "reglas_estructurales": [{"articulo": "74", "regla": "Visita excluyente",
                                      "consecuencia": "No asistir elimina", "severidad": "critica"}],
            "supuestos_competencia": {"n_competidores": [2, 4], "precio_rival_pct": [0.80, 1.0]},
            "plan_accion": [{"prioridad": "🔴 0", "accion": "Asistir a la visita", "asegura": "Habilita todo",
                             "costo": "≈ $0", "plazo": "04-AGO", "responsable": "Pablo"}]},
    }


def _lic(codigo="A", **kw):
    base = {"codigo": codigo, "nombre": "Container doble piso", "organismo": "Muni Y",
            "region": "Ñuble", "monto_estimado": 30000000}
    base.update(kw)
    with db.conn() as c:
        db.upsert_licitacion(c, base)


def _capa_c(codigo, presupuesto=28000000, ex=None):
    ex = ex if ex is not None else {"criterios_evaluacion": [{"nombre": "Precio", "peso_pct": 60}],
                                    "requisitos_clave": ["Visita obligatoria"],
                                    "garantias": [{"tipo": "Seriedad", "monto_o_pct": "3%"}]}
    with db.conn() as c:
        c.execute("INSERT INTO analisis_bases (codigo, ts, resumen, presupuesto_clp, plazo_dias, "
                  "json_extraccion) VALUES (?,?,?,?,?,?)",
                  (codigo, "2026-08-01", "Bases claras.", presupuesto, 45, json.dumps(ex)))


def _cubicacion(codigo, cantidad=2, precio=3051000):
    with db.conn() as c:
        c.execute("INSERT INTO cubicaciones (proyecto, codigo, origen, fecha) VALUES (?,?,?,?)",
                  (codigo, codigo, "ia", "2026-08-01"))
        cid = c.execute("SELECT id FROM cubicaciones WHERE codigo=?", (codigo,)).fetchone()[0]
        c.execute("INSERT INTO cubicacion_items (cubicacion_id, partida, grupo, descripcion, cantidad, "
                  "unidad, precio_unitario, total) VALUES (?,?,?,?,?,?,?,?)",
                  (cid, "1.1", "material", "Container oficina modular", cantidad, "un", precio,
                   cantidad * precio))


def _wb(b):
    return load_workbook(io.BytesIO(b))


def _texto(ws):
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)


# --------------------------------------------------------------------------------- casos base
def test_codigo_inexistente_da_error_no_crashea(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    r = informe.generar_xlsx("NO-EXISTE", enriquecer=_sin_ia)
    assert isinstance(r, dict) and r.get("error")


def test_licitacion_pelada_igual_genera_las_hojas_base(tmp_path, monkeypatch):
    """El caso más común: sin Capa C, sin cubicación y sin IA. Igual tiene que salir un Excel."""
    _setup(tmp_path, monkeypatch)
    _lic("A", nombre="Sala modular", monto_estimado=50000000)
    r = informe.generar("A", enriquecer=_sin_ia)
    assert r["ok"], r.get("error")
    assert r["hojas"] == ["Resumen", "Objeto y Alineación", "Cubicación", "Compras",
                          "Financiero", "Riesgos", "RACI"]


def test_informe_completo_arma_las_nueve_hojas(tmp_path, monkeypatch):
    """Con matriz de evaluación (IA) y censo de competencia aparecen las dos hojas opcionales."""
    _setup(tmp_path, monkeypatch)
    _lic("B")
    _capa_c("B")
    _cubicacion("B")
    with db.conn() as c:      # censo mínimo para que se dispare la hoja Competencia
        c.execute("INSERT INTO competidores (rut, razon_social) VALUES (?,?)", ("1-9", "Rival SpA"))
        c.execute("INSERT INTO participaciones (rut, codigo, resultado, anio) VALUES (?,?,?,?)",
                  ("1-9", "B", "ganada", 2025))
    r = informe.generar("B", enriquecer=_con_ia)
    assert r["hojas"] == ["Resumen", "Objeto y Alineación", "Cubicación", "Compras", "Financiero",
                          "Riesgos", "RACI", "Estrategia Ganar", "Competencia"]
    wb = _wb(r["xlsx"])
    assert "Pago a 60 días" in _texto(wb["Riesgos"])
    assert "Container 20ft" in _texto(wb["Compras"])
    assert "Visita excluyente" in _texto(wb["Estrategia Ganar"])
    assert "Rival SpA" in _texto(wb["Competencia"])


def test_sin_ia_avisa_en_vez_de_mentir(tmp_path, monkeypatch):
    """Si la IA no corre, el Excel NO debe simular que hay análisis: tiene que decirlo."""
    _setup(tmp_path, monkeypatch)
    _lic("C")

    def revienta(*a, **k):
        raise RuntimeError("sin API key")

    r = informe.generar("C", enriquecer=revienta)
    assert r["ok"] and r["con_ia"] is False
    assert any("IA no disponible" in x for x in r["info_faltante"])
    assert "Sin plan de compras" in _texto(_wb(r["xlsx"])["Compras"])


# --------------------------------------------------------------------------------- fórmulas vivas
def test_las_celdas_derivadas_son_formulas_no_valores_pegados(tmp_path, monkeypatch):
    """El punto del Excel sobre el Word: editar Cantidad o Precio recalcula TODO."""
    _setup(tmp_path, monkeypatch)
    _lic("D")
    _capa_c("D")
    _cubicacion("D")
    wb = _wb(informe.generar("D", enriquecer=_sin_ia)["xlsx"])

    cub = wb["Cubicación"]
    fila = next(r for r in cub.iter_rows() if any("Container" in str(c.value or "") for c in r))
    assert str(fila[4].value).startswith("=")              # Subtotal = C*D

    fin = wb["Financiero"]
    formulas = {str(r[0].value).strip(): str(r[1].value) for r in fin.iter_rows()
                if r[0].value and len(r) > 1}
    assert formulas["Costo base directo"].startswith("='Cubicación'!")     # cruza de hoja
    for etiqueta in ("+ sobrecosto", "Costo financiamiento (tasa × capital)", "Precio objetivo",
                     "Costo total cargado"):
        assert formulas[etiqueta].startswith("="), f"{etiqueta} quedó pegado como valor"

    res = wb["Resumen"]
    kpis = [str(r[1].value) for r in res.iter_rows()
            if r[0].value and str(r[0].value).startswith("Costo base directo") and len(r) > 1]
    assert kpis and kpis[0].startswith("='Financiero'!")   # el Resumen mira al Financiero, no copia


def test_escenarios_de_oferta_cuelgan_del_techo_editable(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _lic("E")
    _capa_c("E")
    _cubicacion("E")
    res = _wb(informe.generar("E", enriquecer=_sin_ia)["xlsx"])["Resumen"]
    txt = _texto(res)
    assert "ESCENARIOS DE OFERTA vs. TECHO" in txt
    formulas = [str(c.value) for row in res.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=") and "Financiero" in c.value]
    assert len(formulas) >= 5, "los escenarios deben referenciar al Financiero, no traer números"


def test_sin_presupuesto_no_revienta_el_formato_condicional(tmp_path, monkeypatch):
    """Bug real: sin techo no hay escenarios, y aplicar formato condicional sobre un rango vacío
    (E7:E6) rompía el archivo. Con presupuesto 0 el informe igual tiene que salir."""
    _setup(tmp_path, monkeypatch)
    _lic("F", monto_estimado=None)
    _capa_c("F", presupuesto=None)
    r = informe.generar("F", enriquecer=_sin_ia)
    assert r["ok"]
    assert "Sin presupuesto publicado" in _texto(_wb(r["xlsx"])["Resumen"])


def test_sin_cubicacion_ni_riesgos_los_sum_siguen_siendo_validos(tmp_path, monkeypatch):
    """SUM(E4:E3) es un rango inválido: con listas vacías hay que dejar una fila, no un rango dado vuelta."""
    _setup(tmp_path, monkeypatch)
    _lic("G")
    wb = _wb(informe.generar("G", enriquecer=_sin_ia)["xlsx"])
    for hoja, col in (("Cubicación", 5), ("Riesgos", 6)):
        sums = [str(c.value) for row in wb[hoja].iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=SUM(")]
        assert sums, f"falta el total en {hoja}"
        for s in sums:
            rango = s[len("=SUM("):-1].split(":")
            f1 = int("".join(ch for ch in rango[0] if ch.isdigit()))
            f2 = int("".join(ch for ch in rango[1] if ch.isdigit()))
            assert f2 >= f1, f"rango invertido en {hoja}: {s}"


# --------------------------------------------------------------------------------- guardado
def test_siempre_guarda_en_la_carpeta_de_bases(tmp_path, monkeypatch):
    """Requisito explícito: el Excel queda en la carpeta de la licitación, no solo en la descarga."""
    _setup(tmp_path, monkeypatch)
    madre = tmp_path / "Loveo Construcciones" / "01. Bases"
    (madre / "H").mkdir(parents=True)
    monkeypatch.setenv("LOVEO_BASES_LOCAL", str(madre))
    _lic("H")
    r = informe.generar("H", enriquecer=_sin_ia)
    assert r["ruta"] and (madre / "H" / informe.nombre_archivo("H")).exists()
    assert (madre / "H" / informe.nombre_archivo("H")).read_bytes() == r["xlsx"]


def test_guarda_junto_a_los_documentos_ya_registrados(tmp_path, monkeypatch):
    """Si las bases ya están en disco, el análisis va a ESA carpeta (la que abre Martín),
    aunque LOVEO_BASES_LOCAL apunte a otro lado."""
    _setup(tmp_path, monkeypatch)
    real = tmp_path / "otra_ruta" / "I"
    real.mkdir(parents=True)
    (real / "bases.pdf").write_bytes(b"x")
    _lic("I")
    with db.conn() as c:
        c.execute("INSERT INTO documentos (codigo, nombre_archivo, ruta_local, tipo) VALUES (?,?,?,?)",
                  ("I", "bases.pdf", str(real / "bases.pdf"), "pdf"))
    r = informe.generar("I", enriquecer=_sin_ia)
    assert r["ruta"] == str(real / informe.nombre_archivo("I"))


def test_sin_carpeta_configurada_cae_a_storage(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _lic("J")
    r = informe.generar("J", enriquecer=_sin_ia)
    assert r["ruta"] and str(tmp_path / "storage" / "J") in r["ruta"]


def test_un_disco_que_falla_no_mata_la_descarga(tmp_path, monkeypatch):
    """Guardar es obligatorio, pero si el disco falla el usuario igual se lleva su Excel."""
    _setup(tmp_path, monkeypatch)
    _lic("K")

    def no_se_puede(_codigo):
        raise OSError("OneDrive desconectado")

    monkeypatch.setattr(informe, "carpeta_destino", no_se_puede)
    r = informe.generar("K", enriquecer=_sin_ia)
    assert r["ok"] and r["xlsx"] and r["ruta"] is None and r["aviso"]


# --------------------------------------------------------------------------------- coherencia
def test_el_veredicto_coincide_con_el_modelo_financiero(tmp_path, monkeypatch):
    """El semáforo del Resumen y los factores del consolidado tienen que salir de los MISMOS
    parámetros. Si alguien toca markup/sobrecosto, esto se cae — que es lo que queremos."""
    P = analisis.PARAMS
    assert analisis.FACTOR_SANO == pytest.approx(P["markup_obj"] * (1 + P["sobrecosto"]))
    assert analisis.FACTOR_LIMITE == pytest.approx(P["markup_min"] * (1 + P["sobrecosto"]))
    assert analisis.FACTOR_SANO == pytest.approx(1.885)      # techo neto / 1,885 = costo máx. SANO
    assert analisis.FACTOR_LIMITE == pytest.approx(1.625)    # techo neto / 1,625 = costo máx. LÍMITE


def test_scoring_del_informe_es_el_mismo_del_tablero(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _lic("L")
    sj = {"dimensiones": {d: {"score": 9, "evaluado": True} for d in scoring.PESOS},
          "score_provisional": 108}
    with db.conn() as c:
        db.set_score_provisional(c, "L", 108, json.dumps(sj), False)
    res = _wb(informe.generar("L", enriquecer=_sin_ia)["xlsx"])["Resumen"]
    total = next(r[3].value for r in res.iter_rows()
                 if r[0].value == "TOTAL / 120" and len(r) > 3)
    assert str(total).startswith("=SUM(")           # el total también es fórmula
    assert "GO FUERTE" in _texto(res)               # 9×12 = 108 → GO FUERTE

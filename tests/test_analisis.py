"""
test_analisis.py — Sensor del armado de datos (analisis.py) y del consolidado (consolidado.py).

Lo que protege:
  1. La cubicación se PRECIA desde `precios_referencia` (los 91 precios reales que pagó Loveo),
     no desde lo que el modelo recuerde. Un precio inventado mueve el margen y el GO/NO-GO.
  2. La cubicación curada por el usuario/Valentina MANDA sobre la que proponga la IA.
  3. Un fallo de la IA degrada el informe, no lo mata.
  4. Los parámetros financieros son UNA sola fuente para el informe y para el consolidado.

  python -m pytest tests/test_analisis.py -q
"""
import json

import pytest

import analisis
import consolidado
import cubicacion
import db
import schema_v2
import schema_v3

PRECIO_CONTAINER = 3_051_000
DESC_CONTAINER = "Container oficina modular nuevo con baño simple completo"


def _setup(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DATABASE_URL", "")
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr(db, "STORAGE_DIR", tmp_path / "storage")
    db.init_db()
    schema_v2.migrate()
    schema_v3.migrate()


def _precio(desc, precio, fuente="valentina", fecha="2026-01-01"):
    with db.conn() as c:
        c.execute("INSERT INTO precios_referencia (descripcion, descripcion_norm, unidad, "
                  "precio_unitario, fuente, fecha) VALUES (?,?,?,?,?,?)",
                  (desc, cubicacion._na(desc), "un", precio, fuente, fecha))


def _lic(codigo="A", **kw):
    base = {"codigo": codigo, "nombre": "Container de oficina", "organismo": "Muni X",
            "region": "Maule", "monto_estimado": 30_000_000, "fecha_cierre": "2099-12-31"}
    base.update(kw)
    with db.conn() as c:
        db.upsert_licitacion(c, base)


def _bom(codigo, descripcion, cantidad=2, precio=None):
    """BOM curado SIN precio (el caso real: la IA extrae partidas, el precio lo pone el catálogo)."""
    with db.conn() as c:
        c.execute("INSERT INTO cubicaciones (proyecto, codigo, origen, fecha) VALUES (?,?,?,?)",
                  (codigo, codigo, "ia", "2026-08-01"))
        cid = c.execute("SELECT id FROM cubicaciones WHERE codigo=?", (codigo,)).fetchone()[0]
        c.execute("INSERT INTO cubicacion_items (cubicacion_id, grupo, descripcion, cantidad, unidad, "
                  "precio_unitario) VALUES (?,?,?,?,?,?)",
                  (cid, "material", descripcion, cantidad, "un", precio))


def _nada(*a, **k):
    return {}


# ------------------------------------------------------------------ precios reales
def test_el_catalogo_sale_de_precios_referencia(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, PRECIO_CONTAINER)
    cat = analisis.precios_catalogo()
    assert cat[cubicacion._na(DESC_CONTAINER)]["precio_unitario"] == PRECIO_CONTAINER


def test_el_catalogo_se_queda_con_el_precio_mas_nuevo(tmp_path, monkeypatch):
    """precios_referencia es un HISTORIAL (se insertan puntos, no se pisan): el análisis usa el último."""
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, 2_800_000, fecha="2025-01-01")
    _precio(DESC_CONTAINER, PRECIO_CONTAINER, fecha="2026-01-01")
    cat = analisis.precios_catalogo()
    assert cat[cubicacion._na(DESC_CONTAINER)]["precio_unitario"] == PRECIO_CONTAINER


def test_la_cubicacion_se_precia_desde_el_catalogo_no_se_inventa(tmp_path, monkeypatch):
    """Una partida sin precio se completa con el precio REAL de Loveo, y queda dicho de dónde salió."""
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, PRECIO_CONTAINER)
    _lic("A")
    _bom("A", DESC_CONTAINER, cantidad=2, precio=None)
    D = analisis.datos("A", enriquecer=_nada)
    fila = D["cubicacion"][0]
    assert fila["precio_unitario"] == PRECIO_CONTAINER
    assert "catálogo" in fila["notas"]
    assert not any("sin precio" in x for x in D["info_faltante"])


def test_una_partida_sin_precio_en_el_catalogo_se_reporta(tmp_path, monkeypatch):
    """Honestidad: si no hay precio, no se inventa uno — se avisa en INFO FALTANTE."""
    _setup(tmp_path, monkeypatch)
    _lic("B")
    _bom("B", "Material exótico sin cotizar", cantidad=3, precio=None)
    D = analisis.datos("B", enriquecer=_nada)
    assert D["cubicacion"][0]["precio_unitario"] == 0
    assert any("sin precio" in x for x in D["info_faltante"])


def test_un_precio_ya_cargado_no_se_pisa(tmp_path, monkeypatch):
    """El preciado del pipeline (o de Valentina) manda sobre el catálogo genérico."""
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, PRECIO_CONTAINER)
    _lic("C")
    _bom("C", DESC_CONTAINER, cantidad=1, precio=9_999_999)
    D = analisis.datos("C", enriquecer=_nada)
    assert D["cubicacion"][0]["precio_unitario"] == 9_999_999


# ------------------------------------------------------------------ precedencia IA vs curado
def test_la_cubicacion_curada_le_gana_a_la_de_la_ia(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, PRECIO_CONTAINER)
    _lic("D")
    _bom("D", DESC_CONTAINER, cantidad=2)

    def ia(*a, **k):
        return {"cubicacion": [{"partida": "Lo que se le ocurrió a la IA", "unidad": "un",
                                "cantidad": 99, "precio_unitario": 1, "notas": ""}]}

    D = analisis.datos("D", enriquecer=ia)
    partidas = [x["partida"] for x in D["cubicacion"]]
    assert DESC_CONTAINER in partidas
    assert "Lo que se le ocurrió a la IA" not in partidas


def test_sin_cubicacion_curada_se_usa_la_de_la_ia_pero_preciada(tmp_path, monkeypatch):
    """Cuando no hay BOM, la IA propone partidas — pero el precio lo pone igual el catálogo."""
    _setup(tmp_path, monkeypatch)
    _precio(DESC_CONTAINER, PRECIO_CONTAINER)
    _lic("E")

    def ia(*a, **k):
        return {"cubicacion": [{"partida": DESC_CONTAINER, "unidad": "un", "cantidad": 2,
                                "precio_unitario": 0, "notas": ""}]}

    D = analisis.datos("E", enriquecer=ia)
    assert D["cubicacion"][0]["precio_unitario"] == PRECIO_CONTAINER


# ------------------------------------------------------------------ degradación
def test_si_la_ia_revienta_el_analisis_sigue_en_pie(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _lic("F")

    def revienta(*a, **k):
        raise RuntimeError("sin API key")

    D = analisis.datos("F", enriquecer=revienta)
    assert D["_ia"] is False
    assert D["meta"]["id"] == "F"
    assert any("IA no disponible" in x for x in D["info_faltante"])


def test_los_criterios_de_capa_c_sirven_de_piso_sin_ia(tmp_path, monkeypatch):
    """Aunque no corra la IA, si la Capa C ya extrajo los criterios la matriz se muestra."""
    _setup(tmp_path, monkeypatch)
    _lic("G")
    ex = {"criterios_evaluacion": [{"nombre": "Oferta Económica", "peso_pct": 40},
                                   {"nombre": "Plazo de entrega", "peso_pct": 60}]}
    with db.conn() as c:
        c.execute("INSERT INTO analisis_bases (codigo, ts, presupuesto_clp, json_extraccion) "
                  "VALUES (?,?,?,?)", ("G", "2026-08-01", 20_000_000, json.dumps(ex)))
    D = analisis.datos("G", enriquecer=_nada)
    crit = D["evaluacion"]["criterios"]
    assert [c["tipo"] for c in crit] == ["continuo", "binario"]   # el de precio se detecta solo


def test_codigo_inexistente(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    assert analisis.datos("NO-EXISTE", enriquecer=_nada).get("error")


# ------------------------------------------------------------------ consolidado
def test_el_consolidado_usa_los_mismos_factores_que_el_informe(tmp_path, monkeypatch):
    P = analisis.PARAMS
    assert analisis.FACTOR_SANO == pytest.approx(P["markup_obj"] * (1 + P["sobrecosto"]))
    assert analisis.FACTOR_LIMITE == pytest.approx(P["markup_min"] * (1 + P["sobrecosto"]))


def test_el_consolidado_solo_trae_vigentes(tmp_path, monkeypatch):
    """La base arrastra licitaciones de hace años: el tablero es para decidir esta semana."""
    _setup(tmp_path, monkeypatch)
    _lic("VIEJA", fecha_cierre="2012-05-23")
    _lic("VIGENTE", fecha_cierre="2099-12-31")
    vigentes = [f["codigo"] for f in consolidado.filas()]
    assert vigentes == ["VIGENTE"]
    assert "VIEJA" in [f["codigo"] for f in consolidado.filas(solo_vigentes=False)]


def test_el_consolidado_reporta_lo_que_falta(tmp_path, monkeypatch):
    _setup(tmp_path, monkeypatch)
    _lic("H", monto_estimado=None)
    f = consolidado.filas()[0]
    assert "presupuesto no publicado" in f["faltante"]
    assert "sin cubicación" in f["faltante"]


def test_los_techos_de_costo_del_consolidado_son_formulas(tmp_path, monkeypatch):
    """Editar el costo estimado de una fila tiene que recalcular su veredicto sin regenerar nada."""
    import io

    from openpyxl import load_workbook

    _setup(tmp_path, monkeypatch)
    _lic("I")
    wb = load_workbook(io.BytesIO(consolidado.consolidado_xlsx()))
    ws = wb["Tablero"]
    fila = next(r for r in ws.iter_rows() if r[1].value == "I")
    assert str(fila[6].value).startswith("=")      # costo máx. SANO
    assert str(fila[7].value).startswith("=")      # costo máx. LÍMITE
    assert str(fila[8].value).startswith("=IF(")   # semáforo
    assert "$B$3" in str(fila[6].value)            # cuelga del factor editable, no de un número
